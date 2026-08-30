from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Body, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Text, ForeignKey, Table, event, func
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from typing import Optional
import os, shutil, uuid, smtplib, json, re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
from companies import (
    COMPANIES,
    DEFAULT_COMPANY_ID,
    company_allows_priority,
    default_priority_for_company,
    get_company,
    normalize_company_id,
)
from attempt_models import backfill_attempts, define_attempt_model, ensure_attempts_schema

# Try importing sendgrid, will fail gracefully if not installed
try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail
except ImportError:
    SendGridAPIClient = None
    Mail = None

# F11: Excel generation (graceful fallback to CSV if openpyxl missing)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None

import base64, io, csv as _csv

load_dotenv()

PST = ZoneInfo("America/Los_Angeles")

app = FastAPI()

origins = os.environ.get("ALLOWED_ORIGINS", "http://localhost:3000").split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def rewrite_api_prefix(request: Request, call_next):
    if request.scope["path"].startswith("/api/"):
        request.scope["path"] = request.scope["path"][4:]
    return await call_next(request)


# DATABASE_URL comes from the environment (.env) in production, e.g.
#   postgresql+psycopg2://user:pass@localhost:5432/photo_tracker
# and falls back to a local SQLite file for development.
DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./photo_tracker.db")
if DATABASE_URL.startswith("sqlite"):
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
else:
    # Postgres/other: check_same_thread is SQLite-only; pre-ping avoids stale conns.
    engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

# Public base URL for absolute links (e.g. the "View photo" hyperlinks in Excel
# exports). Overridable via env for other hosts.
PUBLIC_BASE_URL = os.environ.get(
    "PUBLIC_BASE_URL", "https://159-198-79-219.nip.io")


# ── Many-to-many: Pin (Photo) <-> Profile ──
pin_profile = Table(
    "pin_profile",
    Base.metadata,
    Column("photo_id",   Integer, ForeignKey("photos.id",   ondelete="CASCADE"), primary_key=True),
    Column("profile_id", Integer, ForeignKey("profiles.id", ondelete="CASCADE"), primary_key=True),
)


class Profile(Base):
    __tablename__ = "profiles"
    id           = Column(Integer, primary_key=True, index=True)
    name         = Column(String, nullable=False)
    service_type = Column(String, default="standard")
    note         = Column(Text, nullable=True)
    # Standing pay rate (whole dollars) for this profile — summed across
    # non-completed profiles (including brand-new ones) to produce
    # "Total Available Earnings" on the Earnings screen.
    pay_rate     = Column(Integer, nullable=True)

    # Process-serving company slug (see companies.py). Drives allowed priority
    # levels, diligence attempts, payout schedule, and pay-rate copy.
    company      = Column(String, nullable=True)

    # Standing delivery style (Personal / Sub on 1st / …). Copied onto
    # attempts as completion_type so Add Attempt can lock and prefill it.
    delivery_style = Column(String, nullable=True)

    # ── Profile Location: independent of any Attempt/Photo. Settable before
    # any photo is ever uploaded against this profile — see /upload (Photo)
    # for the separate, GPS-captured Attempt location. ──
    status       = Column(String, nullable=True)  # e.g. "awaiting_attempt"
    address      = Column(String, nullable=True)
    city         = Column(String, nullable=True)
    state        = Column(String, nullable=True)
    postal_code  = Column(String, nullable=True)
    latitude     = Column(Float, nullable=True)
    longitude    = Column(Float, nullable=True)
    file_number     = Column(String, nullable=True)
    created_at   = Column(DateTime, default=datetime.utcnow)
    updated_at   = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    photos       = relationship("Photo", secondary=pin_profile, back_populates="profiles")
    attempts     = relationship(
        "Attempt", back_populates="profile",
        cascade="all, delete-orphan",
    )


Attempt = define_attempt_model(Base)


class Photo(Base):
    """Media attached to an Attempt. Attempt-level fields are also mirrored
    here for backward-compatible map/earnings queries."""
    __tablename__ = "photos"
    id        = Column(Integer, primary_key=True, index=True)
    image_url = Column(String)
    timestamp = Column(DateTime, default=datetime.utcnow)
    latitude  = Column(Float)
    longitude = Column(Float)
    zip_code  = Column(String, nullable=True)
    address   = Column(Text,   nullable=True)   # human-readable address
    note      = Column(Text,   nullable=True)
    # per-photo priority category: standard | special | next_day | asap
    category  = Column(String, nullable=True, default="standard")
    # Completion detail fields (mirror the Rockstar service-record email):
    #   completion_type — e.g. Substitute | Personal | Corporate | Posted
    #   served_to       — who the service was served to
    completion_type = Column(String, nullable=True)
    served_to       = Column(String, nullable=True)
    relation_to     = Column(String, nullable=True)  # relation of served_to to profile
    file_number     = Column(String, nullable=True)
    # Whether the attempt was a successful service. Kept in sync with
    # attempt_status for older clients (1 only when status is "successful").
    successful      = Column(Integer, nullable=False, default=0)
    # Attempt outcome: pending | successful | unsuccessful (default pending).
    attempt_status  = Column(String, nullable=False, default="pending")
    # legacy single profile_id kept for backward compat
    profile_id    = Column(Integer, nullable=True)
    is_favorited  = Column(Integer, nullable=False, default=0)

    # FK to Attempt (Profile 1──* Attempt 1──* Photo)
    attempt_id = Column(Integer, ForeignKey("attempts.id", ondelete="CASCADE"),
                        nullable=True, index=True)

    # ── F1: master-pin grouping. NULL group means a standalone pin whose
    #    own id is its group root. Attempts appended to an existing pin
    #    copy that pin's location_group_id (or its id if none yet). ──
    location_group_id = Column(Integer, nullable=True, index=True)

    # ── F6: timestamp integrity ──
    taken_at           = Column(DateTime, nullable=True)  # device capture time
    original_timestamp = Column(DateTime, nullable=True)  # immutable copy of taken_at
    edited_timestamp   = Column(DateTime, nullable=True)  # last manual edit
    created_at         = Column(DateTime, default=datetime.utcnow)  # pin-creation time (anchors 10-min edit window)

    # ── F7: pay rate (whole dollars) ──
    pay_rate = Column(Integer, nullable=True)

    # ── F10: job lifecycle: open | completed | archived ──
    status       = Column(String, nullable=False, default="open")
    completed_at = Column(DateTime, nullable=True)

    # ── F8/F9: ownership for payouts/earnings ──
    user_id = Column(Integer, nullable=True)

    profiles = relationship("Profile", secondary=pin_profile, back_populates="photos")
    attempt  = relationship("Attempt", back_populates="photos")


class User(Base):
    """Lightweight user record (no password wall for the demo — identity only,
    so attempts/payouts/earnings can be attributed)."""
    __tablename__ = "users"
    id         = Column(Integer, primary_key=True, index=True)
    email      = Column(String, unique=True, nullable=False)
    name       = Column(String, nullable=True)
    role       = Column(String, nullable=False, default="field")  # field | admin
    created_at = Column(DateTime, default=datetime.utcnow)


class EmailRecipient(Base):
    """F11: saved export recipients (5–10 per user)."""
    __tablename__ = "email_recipients"
    id      = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, nullable=True)
    label   = Column(String, nullable=True)
    email   = Column(String, nullable=False)


class PinDraft(Base):
    """F5: optional cloud backup of an in-progress pin."""
    __tablename__ = "pin_drafts"
    id         = Column(String, primary_key=True)   # client-generated uuid
    user_id    = Column(Integer, nullable=True)
    payload    = Column(Text, nullable=False)        # JSON blob
    updated_at = Column(DateTime, default=datetime.utcnow)


class TimestampEdit(Base):
    """F6: audit trail for every timestamp change."""
    __tablename__ = "timestamp_edits"
    id        = Column(Integer, primary_key=True, index=True)
    photo_id  = Column(Integer, ForeignKey("photos.id", ondelete="CASCADE"))
    old_value = Column(DateTime, nullable=True)
    new_value = Column(DateTime, nullable=True)
    edited_by = Column(Integer, nullable=True)
    edited_at = Column(DateTime, default=datetime.utcnow)


class StatusHistory(Base):
    """Audit trail for every job-status change (open/in_progress/completed/
    archived), supporting forward, backward and jump transitions + undo."""
    __tablename__ = "status_history"
    id         = Column(Integer, primary_key=True, index=True)
    photo_id   = Column(Integer, ForeignKey("photos.id", ondelete="CASCADE"), index=True)
    old_status = Column(String, nullable=True)
    new_status = Column(String, nullable=False)
    updated_by = Column(Integer, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow)


class PayoutPeriod(Base):
    """F8: finalized pay-period snapshot (live totals are aggregated on demand)."""
    __tablename__ = "payout_periods"
    id           = Column(Integer, primary_key=True, index=True)
    user_id      = Column(Integer, nullable=True)
    period_start = Column(DateTime, nullable=True)
    period_end   = Column(DateTime, nullable=True)
    total_amount = Column(Integer, nullable=True)
    jobs_count   = Column(Integer, nullable=True)
    finalized_at = Column(DateTime, default=datetime.utcnow)


Base.metadata.create_all(bind=engine)


def _ensure_columns():
    """Lightweight migration: add columns introduced after the table was first
    created. create_all() never alters existing tables. Uses dialect-aware DDL
    so the same code works on SQLite (local) and PostgreSQL (production)."""
    from sqlalchemy import text, inspect
    inspector = inspect(engine)
    is_pg = engine.dialect.name == "postgresql"
    # Postgres rejects SQLite's DATETIME; TIMESTAMP is the portable equivalent.
    DT = "TIMESTAMP" if is_pg else "DATETIME"
    existing = {c["name"] for c in inspector.get_columns("photos")}
    # column name -> DDL fragment to add it
    new_cols = {
        "category":           "VARCHAR DEFAULT 'standard'",
        "is_favorited":       "INTEGER NOT NULL DEFAULT 0",
        "location_group_id":  "INTEGER",
        "taken_at":           DT,
        "original_timestamp": DT,
        "edited_timestamp":   DT,
        "pay_rate":           "INTEGER",
        "completion_type":    "VARCHAR",
        "served_to":          "VARCHAR",
        "status":             "VARCHAR NOT NULL DEFAULT 'open'",
        "completed_at":       DT,
        "user_id":            "INTEGER",
        "created_at":         DT,
        "relation_to":        "VARCHAR",
        "file_number":        "VARCHAR",
        "successful":         "INTEGER NOT NULL DEFAULT 0",
        "attempt_status":     "VARCHAR NOT NULL DEFAULT 'pending'",
        "attempt_id":         "INTEGER",
    }
    with engine.connect() as conn:
        for col, ddl in new_cols.items():
            if col not in existing:
                try:
                    conn.execute(text(f"ALTER TABLE photos ADD COLUMN {col} {ddl}"))
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    print(f"[migrate] skip photos.{col}: {e}")
        # Backfill timestamp-integrity + grouping columns for legacy rows
        conn.execute(text(
            "UPDATE photos SET taken_at = timestamp WHERE taken_at IS NULL"))
        conn.execute(text(
            "UPDATE photos SET original_timestamp = timestamp WHERE original_timestamp IS NULL"))
        conn.execute(text(
            "UPDATE photos SET location_group_id = id WHERE location_group_id IS NULL"))
        conn.execute(text(
            "UPDATE photos SET status = 'in_progress' WHERE status IS NULL"))
        # Backfill completed_at for archived jobs that were archived directly
        # (before the fix that stamps completed_at on archive transitions).
        conn.execute(text(
            "UPDATE photos SET completed_at = COALESCE(edited_timestamp, taken_at, timestamp) "
            "WHERE status IN ('archived', 'completed') AND completed_at IS NULL"))
        conn.commit()
        # Legacy rows: anchor the edit window to their capture time so old pins
        # stay locked (new rows get created_at = insertion time).
        conn.execute(text(
            "UPDATE photos SET created_at = COALESCE(original_timestamp, timestamp) "
            "WHERE created_at IS NULL"))
        # Backfill attempt_status from legacy successful flag when missing.
        try:
            conn.execute(text(
                "UPDATE photos SET attempt_status = CASE "
                "WHEN successful = 1 THEN 'successful' "
                "WHEN successful = 0 THEN 'unsuccessful' "
                "ELSE 'pending' END "
                "WHERE attempt_status IS NULL OR attempt_status = ''"))
            conn.commit()
        except Exception as e:
            conn.rollback()
            print(f"[migrate] skip attempt_status backfill: {e}")
        conn.commit()

        # profiles table: standing per-profile pay rate.
        existing_profiles = {c["name"] for c in inspector.get_columns("profiles")}
        if "pay_rate" not in existing_profiles:
            try:
                conn.execute(text("ALTER TABLE profiles ADD COLUMN pay_rate INTEGER"))
                conn.commit()
            except Exception as e:
                conn.rollback()
                print(f"[migrate] skip profiles.pay_rate: {e}")

        # profiles table: Profile Location + status, independent of any Photo/
        # Attempt. NULL is the correct default for existing profiles — never
        # backfilled from a photo's location (see Profile Location ticket).
        profile_new_cols = {
            "status":       "VARCHAR",
            "address":      "VARCHAR",
            "city":         "VARCHAR",
            "state":        "VARCHAR",
            "postal_code":  "VARCHAR",
            "latitude":     "FLOAT" if not is_pg else "DOUBLE PRECISION",
            "longitude":    "FLOAT" if not is_pg else "DOUBLE PRECISION",
            "created_at":   DT,
            "updated_at":   DT,
            "company":         "VARCHAR",
            "delivery_style":  "VARCHAR",
            "file_number":     "VARCHAR",
        }
        for col, ddl in profile_new_cols.items():
            if col not in existing_profiles:
                try:
                    conn.execute(text(f"ALTER TABLE profiles ADD COLUMN {col} {ddl}"))
                    conn.commit()
                except Exception as e:
                    conn.rollback()
                    print(f"[migrate] skip profiles.{col}: {e}")


_ensure_columns()
# Create attempts table + backfill Profile→Attempt→Photo from legacy photos.
try:
    ensure_attempts_schema(engine, Base, Attempt, Photo, pin_profile)
    _bf_db = SessionLocal()
    try:
        n = backfill_attempts(_bf_db, Attempt, Photo, pin_profile)
        if n:
            print(f"[migrate] backfilled {n} attempt(s) from photos")
    finally:
        _bf_db.close()
except Exception as e:
    print(f"[migrate] attempts backfill skipped: {e}")


def _to_pst_iso(ts):
    if not ts:
        return None
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=timezone.utc)
    return ts.astimezone(PST).isoformat()


def _pst_naive_to_utc(dt):
    """Interpret a naive datetime (as sent by the client's local date/time
    pickers) as PST wall-clock time and convert to naive UTC, matching how
    Photo.timestamp is stored. Without this, filter comparisons are off by
    the PST/UTC offset and silently exclude/include the wrong rows."""
    if dt is None:
        return None
    localized = dt.replace(tzinfo=PST)
    return localized.astimezone(timezone.utc).replace(tzinfo=None)


def _haversine_ft(lat1, lon1, lat2, lon2):
    """Great-circle distance in feet between two lat/lng points."""
    from math import radians, sin, cos, sqrt, atan2
    R_ft = 20925524.9  # Earth radius in feet
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return R_ft * 2 * atan2(sqrt(a), sqrt(1 - a))


def _normalize_attempt_status(ph):
    """Resolve attempt_status from the column or legacy successful flag."""
    raw = (getattr(ph, "attempt_status", None) or "").strip().lower()
    if raw in ("pending", "successful", "unsuccessful"):
        return raw
    if ph.successful is None:
        return "pending"
    return "successful" if ph.successful else "unsuccessful"


def _attempt_is_successful(ph):
    return _normalize_attempt_status(ph) == "successful"


def _resolve_attempt_status_input(attempt_status, successful):
    """Prefer explicit attempt_status; fall back to legacy successful bool."""
    raw = (attempt_status or "").strip().lower() if attempt_status else ""
    if raw in ("pending", "successful", "unsuccessful"):
        return raw
    if successful is True:
        return "successful"
    if successful is False:
        return "unsuccessful"
    return "pending"


def _photo_dict(ph):
    profiles = [{"id": p.id, "name": p.name, "service_type": p.service_type} for p in ph.profiles]
    primary = profiles[0] if profiles else None
    return {
        "id":                 ph.id,
        "image_url":          ph.image_url,
        "timestamp":          _to_pst_iso(ph.timestamp),
        "taken_at":           _to_pst_iso(ph.taken_at),
        "original_timestamp": _to_pst_iso(ph.original_timestamp),
        "edited_timestamp":   _to_pst_iso(ph.edited_timestamp),
        "created_at":         _to_pst_iso(ph.created_at),
        "latitude":           ph.latitude,
        "longitude":          ph.longitude,
        "zip_code":           ph.zip_code,
        "address":            ph.address,
        "note":               ph.note,
        "category":           ph.category or "standard",
        "completion_type":    ph.completion_type,
        "served_to":          ph.served_to,
        "relation_to":        ph.relation_to,
        "file_number":        _inherited_file_number(
            (ph.profiles[0] if getattr(ph, "profiles", None) else None),
            ph.file_number,
        ),
        "successful":         _attempt_is_successful(ph),
        "attempt_status":     _normalize_attempt_status(ph),
        "pay_rate":           ph.pay_rate,
        "status":             ph.status or "open",
        "completed_at":       _to_pst_iso(ph.completed_at),
        "user_id":            ph.user_id,
        "location_group_id":  ph.location_group_id or ph.id,
        "profile_id":         primary["id"]   if primary else ph.profile_id,
        "profile_name":       primary["name"] if primary else "Unknown",
        "service_type":       primary["service_type"] if primary else "standard",
        "profiles":           profiles,
        "is_favorited":       bool(ph.is_favorited),
        "attempt_id":         getattr(ph, "attempt_id", None),
    }


def _attempt_dict(att):
    """Serialize an Attempt with nested photos (media)."""
    photos = list(att.photos or [])
    primary = photos[0] if photos else None
    profile = att.profile
    return {
        "id":                 att.id,
        "profile_id":         att.profile_id,
        "profile_name":       profile.name if profile else "Unknown",
        "service_type":       profile.service_type if profile else "standard",
        "latitude":           att.latitude,
        "longitude":          att.longitude,
        "zip_code":           att.zip_code,
        "address":            att.address,
        "note":               att.note,
        "category":           att.category or "standard",
        "completion_type":    att.completion_type,
        "served_to":          att.served_to,
        "relation_to":        att.relation_to,
        "file_number":        _inherited_file_number(profile, att.file_number),
        "successful":         (att.attempt_status or "") == "successful",
        "attempt_status":     att.attempt_status or "pending",
        "pay_rate":           att.pay_rate,
        "status":             att.status or "open",
        "completed_at":       _to_pst_iso(att.completed_at),
        "user_id":            att.user_id,
        "location_group_id":  att.location_group_id or att.id,
        "timestamp":          _to_pst_iso(att.timestamp),
        "taken_at":           _to_pst_iso(att.taken_at),
        "created_at":         _to_pst_iso(att.created_at),
        # Compat: treat primary photo as the "photo" surface for old clients.
        "image_url":          primary.image_url if primary else None,
        "photo_id":           primary.id if primary else None,
        "photos": [
            {
                "id": primary_ph.id,
                "image_url": primary_ph.image_url,
                "taken_at": _to_pst_iso(primary_ph.taken_at),
                "is_favorited": bool(primary_ph.is_favorited),
            }
            for primary_ph in photos
        ],
        "photo_count": len(photos),
    }


def seed_data():
    # No static seed data — app starts clean
    pass


seed_data()


# ─── PROFILE ROUTES ───────────────────────────────────────────────────────────

def _profile_dict(p):
    """Shared response shape for all three profile endpoints. `attempts_count`
    is the number of Photos (Attempts) logged against this profile — reused
    by the client to decide whether "Awaiting Attempt" still applies and
    whether to show Profile Location vs. the Attempts list."""
    company_id = p.company or DEFAULT_COMPANY_ID
    company = get_company(company_id)
    return {
        "id": p.id,
        "name": p.name,
        "service_type": p.service_type,
        "note": p.note,
        "pay_rate": p.pay_rate,
        "company": company_id,
        "delivery_style": getattr(p, "delivery_style", None),
        "file_number": getattr(p, "file_number", None),
        "company_name": company["name"] if company else None,
        "status": _norm_profile_status(p.status),
        "address": p.address,
        "city": p.city,
        "state": p.state,
        "postal_code": p.postal_code,
        "latitude": p.latitude,
        "longitude": p.longitude,
        "attempts_count": len(p.attempts) if getattr(p, "attempts", None) is not None
            else len(p.photos),
    }


def _is_absent_file_number(value):
    v = (value or "").strip()
    return not v or v.upper() == "N/A"


def _inherited_file_number(profile, stored=None):
    """File number is created on the profile; attempts and photos inherit it."""
    pfn = (getattr(profile, "file_number", None) or "").strip() if profile else ""
    if pfn and pfn.upper() != "N/A":
        return pfn
    s = (stored or "").strip()
    if s:
        return s
    return pfn or None


def _copy_file_number_to_attempts(db, profile):
    """Keep attempt/photo rows in sync with the profile file number."""
    fn = _inherited_file_number(profile)
    if _is_absent_file_number(fn):
        return
    for att in db.query(Attempt).filter(Attempt.profile_id == profile.id).all():
        att.file_number = fn
    for ph in list(getattr(profile, "photos", None) or []):
        ph.file_number = fn


def _file_number_in_use(db, file_number, exclude_profile_id=None):
    """True when another profile (or its attempts) already has this file number.
    N/A and blank are not unique."""
    fn = (file_number or "").strip()
    if not fn or fn.upper() == "N/A":
        return False
    key = fn.lower()
    q = db.query(Profile).filter(func.lower(Profile.file_number) == key)
    if exclude_profile_id is not None:
        q = q.filter(Profile.id != exclude_profile_id)
    if q.first() is not None:
        return True
    aq = db.query(Attempt).filter(func.lower(Attempt.file_number) == key)
    if exclude_profile_id is not None:
        aq = aq.filter(Attempt.profile_id != exclude_profile_id)
    return aq.first() is not None


_PROFILE_STATUS_RANK = {
    "pending": 0,
    "in_progress": 1,
    "completed": 2,
    "archived": 3,
}


def _norm_profile_status(raw):
    """Canonical profile lifecycle: pending | in_progress | completed | archived."""
    s = (raw or "").strip().lower()
    if s in ("awaiting_attempt", "open", ""):
        return "pending"
    if s == "in-progress":
        return "in_progress"
    if s in _PROFILE_STATUS_RANK:
        return s
    return "pending"


def _attempt_outcome(att):
    raw = (getattr(att, "attempt_status", None) or "").strip().lower()
    if raw in ("pending", "successful", "unsuccessful"):
        return raw
    if getattr(att, "successful", 0):
        return "successful"
    return "pending"


def _compute_profile_activity_status(db, profile):
    """pending | in_progress | completed from attempts. Never returns archived."""
    attempts = db.query(Attempt).filter(Attempt.profile_id == profile.id).all()
    if not attempts:
        return "pending"
    if any(_attempt_outcome(a) == "successful" for a in attempts):
        return "completed"
    company = get_company(profile.company or DEFAULT_COMPANY_ID) or {}
    diligence = int(company.get("attempts_for_diligence") or 5)
    unsuccessful = sum(
        1 for a in attempts if _attempt_outcome(a) == "unsuccessful"
    )
    if unsuccessful >= diligence:
        return "completed"
    return "in_progress"


def _advance_profile_status(db, profile):
    """One-way lifecycle. Archived is never changed automatically."""
    current = _norm_profile_status(profile.status)
    if current == "archived":
        if (profile.status or "").strip().lower() != "archived":
            profile.status = "archived"
        return current
    derived = _compute_profile_activity_status(db, profile)
    target = derived
    if _PROFILE_STATUS_RANK[derived] < _PROFILE_STATUS_RANK[current]:
        target = current
    if (profile.status or "") != target:
        profile.status = target
        profile.updated_at = datetime.utcnow()
    return target


def _profile_blocks_new_attempts(profile):
    return _norm_profile_status(profile.status) in ("completed", "archived")


def _parse_profile_location(data, profile):
    """Applies address/city/state/postal_code/lat/lng/status fields from
    `data` onto `profile` in place, when present. Independent of any Photo/
    Attempt — this is the ONLY place (besides create_profile) that ever
    writes Profile Location."""
    for field in ("address", "city", "state", "postal_code"):
        if field in data:
            value = data[field]
            setattr(profile, field, value.strip() if isinstance(value, str) else value)

    for field in ("latitude", "longitude"):
        if field in data:
            value = data[field]
            if value is None:
                setattr(profile, field, None)
                continue
            try:
                value = float(value)
            except (TypeError, ValueError):
                raise HTTPException(status_code=422, detail=f"{field} must be a number")
            if field == "latitude" and not (-90 <= value <= 90):
                raise HTTPException(status_code=422, detail="latitude must be between -90 and 90")
            if field == "longitude" and not (-180 <= value <= 180):
                raise HTTPException(status_code=422, detail="longitude must be between -180 and 180")
            setattr(profile, field, value)


try:
    _lc_db = SessionLocal()
    try:
        for _p in _lc_db.query(Profile).all():
            _advance_profile_status(_lc_db, _p)
        _lc_db.commit()
    finally:
        _lc_db.close()
except Exception as e:
    print(f"[migrate] profile lifecycle backfill skipped: {e}")


@app.get("/companies")
def list_companies():
    """Hardcoded company catalog — clients use this for selectors and to
    resolve rates / diligence / priority allowlists without embedding drift."""
    return COMPANIES


@app.get("/profiles")
def get_profiles():
    db = SessionLocal()
    profiles = db.query(Profile).all()
    result = [_profile_dict(p) for p in profiles]
    db.close()
    return result


@app.post("/profiles")
async def create_profile(data: dict = Body(...)):
    name = data.get("name", "").strip()
    service_type = data.get("service_type", "standard").strip()

    if not name:
        raise HTTPException(status_code=422, detail="Profile name is required")

    company_id = normalize_company_id(data.get("company"), required=True)
    if company_id is None:
        # Blank → default; invalid non-blank → 422
        if (data.get("company") or "").strip():
            raise HTTPException(status_code=422, detail="Unknown company")
        company_id = DEFAULT_COMPANY_ID

    # Priority categories (+ legacy rush/airport kept for backward compat)
    if service_type not in (
        "standard", "special", "next_day", "asap", "rush", "airport"
    ):
        service_type = "standard"
    if not company_allows_priority(company_id, service_type):
        service_type = default_priority_for_company(company_id)

    pay_rate = data.get("pay_rate")
    if pay_rate is not None:
        try:
            pay_rate = int(round(float(pay_rate)))
        except (TypeError, ValueError):
            raise HTTPException(status_code=422, detail="pay_rate must be a whole dollar number")

    delivery_style = (data.get("delivery_style") or "").strip() or None
    file_number = (data.get("file_number") or "").strip() or None

    db = SessionLocal()
    if _file_number_in_use(db, file_number):
        db.close()
        raise HTTPException(
            status_code=409,
            detail="This file number already exists",
        )
    profile = Profile(
        name=name,
        service_type=service_type,
        pay_rate=pay_rate,
        company=company_id,
        delivery_style=delivery_style,
        file_number=file_number,
        status="pending",
    )
    # Profile Location is optional here — creating a
    # profile never requires an attempt, photo, upload, or GPS capture.
    try:
        _parse_profile_location(data, profile)
    except HTTPException:
        db.close()
        raise
    profile.status = "pending"
    db.add(profile)
    db.commit()
    db.refresh(profile)
    result = _profile_dict(profile)
    db.close()
    return result


@app.patch("/profiles/{profile_id}")
async def update_profile(profile_id: int, data: dict = Body(...)):
    db = SessionLocal()
    profile = db.query(Profile).filter(Profile.id == profile_id).first()
    if not profile:
        db.close()
        raise HTTPException(status_code=404, detail="Profile not found")

    if "name" in data:
        profile.name = data["name"]
    if "service_type" in data:
        profile.service_type = data["service_type"]
    if "note" in data:
        profile.note = data["note"]
    if "company" in data:
        company_id = normalize_company_id(data.get("company"), required=True)
        if company_id is None:
            db.close()
            raise HTTPException(status_code=422, detail="Unknown company")
        profile.company = company_id
        # Drop a legacy service_type that the new company doesn't allow.
        if not company_allows_priority(company_id, profile.service_type):
            profile.service_type = default_priority_for_company(company_id)
    if "pay_rate" in data:
        pay_rate = data["pay_rate"]
        if pay_rate is not None:
            try:
                pay_rate = int(round(float(pay_rate)))
            except (TypeError, ValueError):
                db.close()
                raise HTTPException(status_code=422, detail="pay_rate must be a whole dollar number")
        profile.pay_rate = pay_rate
    if "delivery_style" in data:
        style = data["delivery_style"]
        profile.delivery_style = style.strip() if isinstance(style, str) and style.strip() else None
    if "file_number" in data:
        fn = data["file_number"]
        next_fn = fn.strip() if isinstance(fn, str) and fn.strip() else None
        if _file_number_in_use(db, next_fn, exclude_profile_id=profile.id):
            db.close()
            raise HTTPException(
                status_code=409,
                detail="This file number already exists",
            )
        profile.file_number = next_fn
        _copy_file_number_to_attempts(db, profile)

    if "status" in data:
        requested = _norm_profile_status(data.get("status"))
        current = _norm_profile_status(profile.status)
        if requested == "archived":
            if current not in ("completed", "archived"):
                db.close()
                raise HTTPException(
                    status_code=422,
                    detail="Archive is only allowed after the profile is completed",
                )
            profile.status = "archived"
        elif requested != current:
            db.close()
            raise HTTPException(
                status_code=422,
                detail="Profile status is set automatically from attempts",
            )

    # Profile Location can be changed independently of everything else — this
    # never touches Photo/Attempt rows, so historical Attempt GPS is untouched.
    try:
        _parse_profile_location(data, profile)
    except HTTPException:
        db.close()
        raise

    if _norm_profile_status(profile.status) != "archived":
        _advance_profile_status(db, profile)

    profile.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(profile)
    result = _profile_dict(profile)
    db.close()
    return result


@app.delete("/profiles/{profile_id}")
def delete_profile(profile_id: int):
    db = SessionLocal()
    profile = db.query(Profile).filter(Profile.id == profile_id).first()
    if not profile:
        db.close()
        raise HTTPException(status_code=404, detail="Profile not found")

    # Cascade: delete all photos that belong ONLY to this profile
    # (photos linked to multiple profiles are unlinked, not deleted)
    photos_to_delete = []
    for photo in list(profile.photos):
        if len(photo.profiles) == 1:
            # Only linked to this profile — delete the photo file + record
            photos_to_delete.append(photo)
        else:
            # Linked to other profiles too — just remove the association
            photo.profiles = [p for p in photo.profiles if p.id != profile_id]

    for photo in photos_to_delete:
        filepath = photo.image_url.lstrip("/")
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass
        db.delete(photo)

    db.delete(profile)
    db.commit()
    db.close()
    return {"ok": True}


@app.get("/profiles/{profile_id}/photos")
def get_profile_photos(profile_id: int):
    """Legacy endpoint — returns one photo-shaped row per Attempt (primary
    image) so older clients keep working. Prefer GET /profiles/{id}/attempts."""
    db = SessionLocal()
    profile = db.query(Profile).filter(Profile.id == profile_id).first()
    if not profile:
        db.close()
        raise HTTPException(status_code=404, detail="Profile not found")
    attempts = (
        db.query(Attempt)
        .filter(Attempt.profile_id == profile_id)
        .order_by(Attempt.id.desc())
        .all()
    )
    # Newest first by taken_at when present.
    attempts.sort(
        key=lambda a: a.taken_at or a.timestamp or datetime.min,
        reverse=True,
    )
    photos_out = []
    if attempts:
        for att in attempts:
            d = _attempt_dict(att)
            # Shape like _photo_dict for list UIs that still expect photos.
            photos_out.append({
                "id": d.get("photo_id") or d["id"],
                "attempt_id": d["id"],
                "image_url": d.get("image_url"),
                "timestamp": d.get("timestamp"),
                "taken_at": d.get("taken_at"),
                "latitude": d.get("latitude"),
                "longitude": d.get("longitude"),
                "zip_code": d.get("zip_code"),
                "address": d.get("address"),
                "note": d.get("note"),
                "category": d.get("category"),
                "completion_type": d.get("completion_type"),
                "served_to": d.get("served_to"),
                "relation_to": d.get("relation_to"),
                "file_number": d.get("file_number"),
                "successful": d.get("successful"),
                "attempt_status": d.get("attempt_status"),
                "pay_rate": d.get("pay_rate"),
                "status": d.get("status"),
                "completed_at": d.get("completed_at"),
                "location_group_id": d.get("location_group_id"),
                "profile_id": d.get("profile_id"),
                "profile_name": d.get("profile_name"),
                "service_type": d.get("service_type"),
                "photos": d.get("photos"),
                "photo_count": d.get("photo_count"),
                "is_favorited": False,
            })
    else:
        # Fallback before backfill finished
        photos_out = [_photo_dict(ph) for ph in profile.photos]
    result = {
        "profile": {"id": profile.id, "name": profile.name, "service_type": profile.service_type},
        "photos":  photos_out,
    }
    db.close()
    return result


@app.get("/profiles/{profile_id}/attempts")
def get_profile_attempts(profile_id: int):
    """Profile → Attempts (newest first), each with nested photos."""
    db = SessionLocal()
    profile = db.query(Profile).filter(Profile.id == profile_id).first()
    if not profile:
        db.close()
        raise HTTPException(status_code=404, detail="Profile not found")
    attempts = (
        db.query(Attempt)
        .filter(Attempt.profile_id == profile_id)
        .order_by(Attempt.id.desc())
        .all()
    )
    attempts.sort(
        key=lambda a: a.taken_at or a.timestamp or datetime.min,
        reverse=True,
    )
    result = {
        "profile": _profile_dict(profile),
        "attempts": [_attempt_dict(a) for a in attempts],
    }
    db.close()
    return result


@app.post("/attempts/{attempt_id}/duplicate")
def duplicate_attempt(attempt_id: int, data: dict = Body(...)):
    """Duplicate an attempt onto other nearby profiles.

    Creates distinct Attempt + Photo rows per target profile (independent
    copies — editing one never changes another).
    """
    profile_ids = data.get("profile_ids") or []
    if not isinstance(profile_ids, list) or not profile_ids:
        raise HTTPException(status_code=422, detail="profile_ids required")
    db = SessionLocal()
    source = db.query(Attempt).filter(Attempt.id == attempt_id).first()
    if not source:
        db.close()
        raise HTTPException(status_code=404, detail="Attempt not found")
    created = []
    for raw_pid in profile_ids:
        try:
            pid = int(raw_pid)
        except (TypeError, ValueError):
            continue
        if pid == source.profile_id:
            continue
        profile = db.query(Profile).filter(Profile.id == pid).first()
        if not profile:
            continue
        _advance_profile_status(db, profile)
        if _profile_blocks_new_attempts(profile):
            continue
        clone = Attempt(
            profile_id=pid,
            latitude=source.latitude,
            longitude=source.longitude,
            zip_code=source.zip_code,
            address=source.address,
            note=source.note,
            category=source.category,
            completion_type=source.completion_type,
            served_to=source.served_to,
            relation_to=source.relation_to,
            file_number=_inherited_file_number(profile, source.file_number),
            successful=source.successful,
            attempt_status=source.attempt_status,
            pay_rate=source.pay_rate,
            status=source.status or "in_progress",
            completed_at=None,
            user_id=source.user_id,
            location_group_id=None,
            taken_at=source.taken_at,
            original_timestamp=source.original_timestamp or source.taken_at,
            timestamp=source.timestamp or datetime.utcnow(),
            created_at=datetime.utcnow(),
        )
        db.add(clone)
        db.flush()
        clone.location_group_id = clone.id
        for ph in list(source.photos or []):
            # Copy image file so each attempt owns independent media.
            src_path = (ph.image_url or "").lstrip("/")
            new_url = ph.image_url
            if src_path and os.path.exists(src_path):
                ext = os.path.splitext(src_path)[1] or ".jpg"
                filename = f"{uuid.uuid4()}{ext}"
                dest = os.path.join(UPLOAD_DIR, filename)
                try:
                    shutil.copy2(src_path, dest)
                    new_url = f"/uploads/{filename}"
                except OSError:
                    new_url = ph.image_url
            new_ph = Photo(
                image_url=new_url,
                timestamp=ph.timestamp or clone.timestamp,
                taken_at=ph.taken_at or clone.taken_at,
                original_timestamp=ph.original_timestamp or clone.taken_at,
                latitude=clone.latitude,
                longitude=clone.longitude,
                zip_code=clone.zip_code,
                address=clone.address,
                note=clone.note,
                category=clone.category,
                completion_type=clone.completion_type,
                served_to=clone.served_to,
                relation_to=clone.relation_to,
                file_number=clone.file_number,
                successful=clone.successful,
                attempt_status=clone.attempt_status,
                pay_rate=clone.pay_rate,
                user_id=clone.user_id,
                status=clone.status,
                profile_id=pid,
                attempt_id=clone.id,
                location_group_id=clone.id,
            )
            new_ph.profiles = [profile]
            db.add(new_ph)
        created.append(clone)
        _advance_profile_status(db, profile)
    db.commit()
    out = []
    for a in created:
        db.refresh(a)
        out.append(_attempt_dict(a))
    db.close()
    return {"ok": True, "duplicated": len(out), "attempts": out}


@app.get("/photos/{photo_id}/watermarked")
def watermarked_photo(photo_id: int, max_w: int = 1600, max_h: int = 1600):
    """Return the photo with its timestamp + geotag watermark burned in. Used
    by the 'View photo' hyperlinks in the Excel exports (full 1600px size, so
    the sheet stays compact while the linked image is still stamped) and by
    inline <img> tags in export emails, which pass a smaller max_w/max_h
    (e.g. 800) — a full-size PNG is too heavy to render reliably inline."""
    from fastapi.responses import Response
    db = SessionLocal()
    ph = db.query(Photo).filter(Photo.id == photo_id).first()
    if not ph or not ph.image_url:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    # Build caption while the session is open so profile names can lazy-load.
    caption = _photo_caption(ph)
    path = ph.image_url.lstrip("/")
    db.close()
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Image file missing")
    stamped = _watermark_photo_png(path, caption, max_w=max_w, max_h=max_h)
    if stamped:
        return Response(content=stamped[0], media_type="image/png")
    with open(path, "rb") as f:                       # fallback: raw image
        return Response(content=f.read(), media_type="image/jpeg")


# ─── PHOTO ROUTES ─────────────────────────────────────────────────────────────

@app.get("/photos")
def get_photos(
    user_lat: Optional[float] = None,
    user_lng: Optional[float] = None,
):
    """Return all photos. When user_lat/user_lng are provided each photo gets
    a distance_mi field and results are sorted nearest-first."""
    db = SessionLocal()
    photos = db.query(Photo).order_by(Photo.id.desc()).all()
    result = []
    for ph in photos:
        d = _photo_dict(ph)
        if user_lat is not None and user_lng is not None and ph.latitude and ph.longitude:
            dist_ft = _haversine_ft(user_lat, user_lng, ph.latitude, ph.longitude)
            d["distance_mi"] = round(dist_ft / 5280, 2)
        else:
            d["distance_mi"] = None
        result.append(d)
    db.close()
    if user_lat is not None and user_lng is not None:
        result.sort(key=lambda d: d["distance_mi"] if d["distance_mi"] is not None else float("inf"))
    return result


def _parse_device_ts(s):
    """Parse a device-supplied ISO timestamp; return None on failure."""
    if not s:
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.replace("Z", "").split("+")[0], fmt)
        except ValueError:
            continue
    return None


@app.post("/upload")
async def upload_photo(
    file:              UploadFile = File(...),
    profile_id:        int        = Form(...),
    latitude:          float      = Form(...),
    longitude:         float      = Form(...),
    zip_code:          str        = Form(""),
    address:           str        = Form(""),
    note:              str        = Form(""),
    category:          str        = Form("standard"),
    completion_type:   str        = Form(""),    # e.g. Substitute | Personal
    served_to:         str        = Form(""),    # who service was served to
    relation_to:       str        = Form(""),    # served_to's relation to the profile
    file_number:       str             = Form(""),    # dispatcher-assigned file number
    successful:        Optional[bool]  = Form(None),  # legacy; prefer attempt_status
    attempt_status:    Optional[str]   = Form(None),  # pending|successful|unsuccessful
    taken_at:          str             = Form(""),    # F6: device capture time (ISO)
    pay_rate:          str        = Form(""),     # F7: whole dollars
    user_id:           str        = Form(""),     # F8/F9: attribution
    location_group_id: str        = Form(""),     # F1: append to existing master pin
    attempt_id:        str        = Form(""),     # attach photo to existing Attempt
):
    # Validate coordinates
    if not (-90 <= latitude <= 90):
        raise HTTPException(status_code=422, detail="Latitude must be between -90 and 90")
    if not (-180 <= longitude <= 180):
        raise HTTPException(status_code=422, detail="Longitude must be between -180 and 180")
    ext      = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"{uuid.uuid4()}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    db = SessionLocal()
    profile = db.query(Profile).filter(Profile.id == profile_id).first()
    if not profile:
        db.close()
        raise HTTPException(status_code=404, detail="Profile not found")

    resolved_file_number = _inherited_file_number(profile, file_number)
    if _is_absent_file_number(resolved_file_number):
        resolved_file_number = (file_number or "").strip() or "N/A"

    # F6: lock timestamp to device capture time when provided; the upload time
    # is no longer authoritative. original_timestamp is the immutable record.
    device_ts = _parse_device_ts(taken_at) or datetime.utcnow()

    # F7: parse whole-dollar pay rate
    pay_val = None
    if pay_rate.strip():
        try:
            pay_val = int(round(float(pay_rate.strip())))
        except ValueError:
            pay_val = None

    uid = int(user_id) if user_id.strip().isdigit() else None
    grp = int(location_group_id) if location_group_id.strip().isdigit() else None
    existing_attempt_id = int(attempt_id) if attempt_id.strip().isdigit() else None

    _advance_profile_status(db, profile)
    if not existing_attempt_id and _profile_blocks_new_attempts(profile):
        db.close()
        try:
            os.remove(filepath)
        except OSError:
            pass
        raise HTTPException(
            status_code=422,
            detail="This profile is completed and cannot receive new attempts",
        )

    if not existing_attempt_id:
        plat = getattr(profile, "latitude", None)
        plng = getattr(profile, "longitude", None)
        if plat is not None and plng is not None and (abs(plat) > 0.0001 or abs(plng) > 0.0001):
            dist_ft = _haversine_ft(latitude, longitude, plat, plng)
            if dist_ft > 200:
                db.close()
                try:
                    os.remove(filepath)
                except OSError:
                    pass
                raise HTTPException(
                    status_code=422,
                    detail=(
                        f"You are {int(round(dist_ft))} ft from this job. "
                        "Attempts must be logged within 200 feet of the profile location."
                    ),
                )

    # New pins always start as in_progress — status only changes to completed
    # or archived when the user explicitly closes them out.
    initial_status = "in_progress"
    initial_completed_at = None

    # Resolve priority against the profile's company allowlist.
    cat = category.strip().lower()
    if cat not in ("standard", "special", "next_day", "asap"):
        cat = "standard"
    company_id = profile.company or DEFAULT_COMPANY_ID
    if not company_allows_priority(company_id, cat):
        cat = default_priority_for_company(company_id)

    attempt_st = _resolve_attempt_status_input(attempt_status, successful)

    # Resolve or create the Attempt (upload session = one Attempt).
    attempt = None
    if existing_attempt_id:
        attempt = db.query(Attempt).filter(
            Attempt.id == existing_attempt_id,
            Attempt.profile_id == profile_id,
        ).first()
        if not attempt:
            db.close()
            raise HTTPException(status_code=404, detail="Attempt not found")
        if _is_absent_file_number(attempt.file_number) and resolved_file_number:
            attempt.file_number = resolved_file_number
    else:
        attempt = Attempt(
            profile_id=profile_id,
            latitude=latitude,
            longitude=longitude,
            zip_code=zip_code.strip() or None,
            address=address.strip() or None,
            note=note.strip() or None,
            category=cat,
            completion_type=(getattr(profile, "delivery_style", None) or "").strip() or None,
            served_to=served_to.strip() or None,
            relation_to=relation_to.strip() or None,
            file_number=resolved_file_number,
            successful=1 if attempt_st == "successful" else 0,
            attempt_status=attempt_st,
            pay_rate=pay_val,
            user_id=uid,
            status=initial_status,
            completed_at=initial_completed_at,
            taken_at=device_ts,
            original_timestamp=device_ts,
            timestamp=device_ts,
            created_at=datetime.utcnow(),
            location_group_id=grp,
        )
        db.add(attempt)
        db.flush()
        if not attempt.location_group_id:
            attempt.location_group_id = attempt.id

    photo = Photo(
        image_url          = f"/uploads/{filename}",
        timestamp          = device_ts,
        taken_at           = device_ts,
        original_timestamp = device_ts,
        latitude           = attempt.latitude,
        longitude          = attempt.longitude,
        zip_code           = attempt.zip_code,
        address            = attempt.address,
        note               = attempt.note,
        category           = attempt.category,
        completion_type    = attempt.completion_type,
        served_to          = attempt.served_to,
        relation_to        = attempt.relation_to,
        file_number        = _inherited_file_number(profile, attempt.file_number),
        successful         = attempt.successful,
        attempt_status     = attempt.attempt_status,
        pay_rate           = attempt.pay_rate,
        user_id            = uid,
        status             = attempt.status or initial_status,
        completed_at       = attempt.completed_at,
        profile_id         = profile_id,
        attempt_id         = attempt.id,
        location_group_id  = attempt.location_group_id or attempt.id,
    )
    photo.profiles = [profile]
    db.add(photo)
    db.flush()
    _advance_profile_status(db, profile)
    db.commit()
    db.refresh(photo)
    db.refresh(attempt)
    result = _photo_dict(photo)
    result["attempt_id"] = attempt.id
    result["attempt"] = _attempt_dict(attempt)
    db.close()
    return result


@app.patch("/photos/{photo_id}/image")
async def replace_photo_image(photo_id: int, file: UploadFile = File(...)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    # Delete old file
    old_path = photo.image_url.lstrip("/")
    if os.path.exists(old_path):
        os.remove(old_path)
    # Save new file
    ext      = os.path.splitext(file.filename)[1] or ".jpg"
    filename = f"{uuid.uuid4()}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    photo.image_url = f"/uploads/{filename}"
    db.commit()
    result = _photo_dict(photo)
    db.close()
    return result


@app.patch("/photos/{photo_id}/location")
async def update_photo_location(photo_id: int, data: dict = Body(None), latitude: float = Form(None), longitude: float = Form(None)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    
    if data:
        photo.latitude  = data.get("latitude",  photo.latitude)
        photo.longitude = data.get("longitude", photo.longitude)
    else:
        if latitude is not None:
            photo.latitude = latitude
        if longitude is not None:
            photo.longitude = longitude
    
    db.commit()
    db.close()
    return {"ok": True}


@app.patch("/photos/{photo_id}/note")
async def update_photo_note(photo_id: int, data: dict = Body(None), note: str = Form(None)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    
    if data:
        photo.note = data.get("note", photo.note)
    else:
        if note is not None:
            photo.note = note
    
    db.commit()
    db.close()
    return {"ok": True}


@app.patch("/photos/{photo_id}/zip")
async def update_photo_zip(photo_id: int, data: dict = Body(None), zip_code: str = Form(None)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    
    if data:
        photo.zip_code = data.get("zip_code", photo.zip_code)
    else:
        if zip_code is not None:
            photo.zip_code = zip_code
    
    db.commit()
    db.close()
    return {"ok": True}


@app.patch("/photos/{photo_id}/address")
async def update_photo_address(photo_id: int, data: dict = Body(...)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    photo.address  = data.get("address",  photo.address)
    photo.zip_code = data.get("zip_code", photo.zip_code)
    db.commit()
    db.close()
    return {"ok": True}


@app.patch("/photos/{photo_id}/category")
async def update_photo_category(photo_id: int, data: dict = Body(...)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    cat = (data.get("category") or "standard").strip().lower()
    if cat not in ("standard", "special", "next_day", "asap"):
        cat = "standard"
    photo.category = cat
    db.commit()
    db.close()
    return {"ok": True, "category": cat}


@app.patch("/photos/{photo_id}/favorite")
async def toggle_favorite(photo_id: int):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    new_val = 0 if photo.is_favorited else 1
    photo.is_favorited = new_val
    db.commit()
    db.close()
    return {"ok": True, "is_favorited": bool(new_val)}


@app.patch("/photos/{photo_id}/profiles")
async def update_photo_profiles(photo_id: int, data: dict = Body(None), profile_ids: str = Form(None)):
    """Add or replace profiles associated with a pin."""
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    
    if data:
        profile_ids_list = data.get("profile_ids", [])
    else:
        # Parse comma-separated profile IDs from form
        profile_ids_list = [int(x.strip()) for x in profile_ids.split(",") if x.strip()] if profile_ids else []
    
    profiles = db.query(Profile).filter(Profile.id.in_(profile_ids_list)).all()
    photo.profiles = profiles
    db.commit()
    db.close()
    return {"ok": True}


@app.delete("/photos/{photo_id}")
def delete_photo(photo_id: int):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    filepath = photo.image_url.lstrip("/")
    if os.path.exists(filepath):
        os.remove(filepath)
    db.delete(photo)
    db.commit()
    db.close()
    return {"ok": True}


# ─── LOG ROUTE ────────────────────────────────────────────────────────────────

@app.get("/log")
def get_log(
    date:       Optional[str] = None,
    start_time: Optional[str] = None,   # ISO datetime or "YYYY-MM-DD HH:MM"
    end_time:   Optional[str] = None,   # ISO datetime or "YYYY-MM-DD HH:MM"
    zip_code:   Optional[str] = None,
    status:     Optional[str] = None,
    search:     Optional[str] = None,
):
    db = SessionLocal()
    query = db.query(Photo)

    # ── Time range filtering ──────────────────────────────────────────────
    if start_time or end_time:
        # Explicit time range takes priority over date-only filter
        def _parse_dt(s):
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s, fmt)
                except ValueError:
                    continue
            return None

        if start_time:
            dt_start = _parse_dt(start_time)
            if dt_start:
                query = query.filter(Photo.timestamp >= _pst_naive_to_utc(dt_start))
        if end_time:
            dt_end = _parse_dt(end_time)
            if dt_end:
                # If only date provided (no time), extend to end of that day
                if "T" not in end_time and " " not in end_time.strip():
                    dt_end = dt_end.replace(hour=23, minute=59, second=59)
                query = query.filter(Photo.timestamp <= _pst_naive_to_utc(dt_end))
    elif date:
        try:
            d = datetime.strptime(date, "%Y-%m-%d")
            query = query.filter(
                Photo.timestamp >= _pst_naive_to_utc(d.replace(hour=0, minute=0, second=0)),
                Photo.timestamp <= _pst_naive_to_utc(d.replace(hour=23, minute=59, second=59)),
            )
        except ValueError:
            pass

    if zip_code:
        query = query.filter(Photo.zip_code.ilike(f"%{zip_code}%"))

    photos = query.order_by(Photo.timestamp.desc()).all()
    result = []
    for ph in photos:
        d = _photo_dict(ph)
        # filter by service status after loading (needs profile data)
        if status and d["service_type"] != status:
            continue
        # filter by note text search — also matches profile names, zip, and address
        if search:
            s = search.lower()
            note_match    = ph.note    and s in ph.note.lower()
            address_match = ph.address and s in ph.address.lower()
            profile_match = any(s in p["name"].lower() for p in d["profiles"])
            zip_match     = ph.zip_code and s in ph.zip_code.lower()
            if not (note_match or address_match or profile_match or zip_match):
                continue
        result.append(d)

    db.close()
    return result


# ─── EMAIL EXPORT ─────────────────────────────────────────────────────────────

@app.post("/export/email")
async def export_log_email(request: Request):
    """
    Send a log export to an email address.
    Supports both JSON body (mobile) and Form data (web)
    """
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        data = await request.json()
        to_email = data.get("to", "").strip()
        records_list = data.get("records", [])
    else:
        form = await request.form()
        to_email = (form.get("to") or "").strip()
        try:
            records_list = json.loads(form.get("records", "[]"))
        except Exception:
            records_list = []

    if not to_email:
        raise HTTPException(status_code=422, detail="Email address required")

    # Category display metadata: label + colour
    _CATEGORY_META = {
        "standard": ("Standard", "#059669"),
        "special":  ("Special",  "#EA580C"),
        "next_day": ("Next Day", "#CA8A04"),
        "asap":     ("ASAP",     "#DC2626"),
    }

    # Build HTML table
    rows_html = ""
    for r in records_list:
        cat_label, cat_color = _CATEGORY_META.get(
            (r.get("category") or "standard").lower(),
            _CATEGORY_META["standard"],
        )
        profiles = ", ".join(p["name"] for p in r.get("profiles", [])) or r.get("profile_name", "—")
        lat = r.get('latitude', '')
        lng = r.get('longitude', '')
        coords = f"{lat:.4f}, {lng:.4f}" if isinstance(lat, (int, float)) and isinstance(lng, (int, float)) else "—"
        address = r.get('address') or "—"
        zip_val = r.get('zip_code') or "—"
        # Format address inline with ZIP
        if r.get('address') and r.get('zip_code') and r['zip_code'] not in r['address']:
            address = f"{r['address']}, {r['zip_code']}"
        elif r.get('address'):
            address = r['address']
        rows_html += f"""
        <tr>
          <td>{r.get('timestamp','—')}</td>
          <td>{profiles}</td>
          <td>{r.get('service_type','—').upper()}</td>
          <td><span style="color:{cat_color};font-weight:700;">{cat_label}</span></td>
          <td>{address}</td>
          <td>{coords}</td>
          <td>{r.get('note','—')}</td>
        </tr>"""

    table_html = f"""
    <table border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:100%;font-size:13px;">
      <thead style="background:#f1f5f9;">
        <tr>
          <th>Timestamp (PST)</th><th>Profile(s)</th><th>Status</th>
          <th>Category</th><th>Address</th><th>Coordinates</th><th>Note</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>"""
    html = _pro_email_html(
        f"Please find the GeoTagging CRM activity log you requested below, "
        f"containing {len(records_list)} record(s). All times are shown in PST.",
        table_html=table_html, count=len(records_list),
        closing="A summary of the exported records is included above. If you "
                "have any questions or need anything further, simply reply to "
                "this email.")

    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    
    sg_api_key = os.environ.get("SENDGRID_API_KEY")
    sg_sender = os.environ.get("SENDGRID_SENDER_EMAIL")

    if not sg_api_key and (not smtp_host or not smtp_user):
        # Return the HTML as a download instead
        return {"ok": True, "message": "Email not configured — export data returned", "html": html, "count": len(records_list)}

    try:
        # Use SendGrid if configured
        if sg_api_key and SendGridAPIClient and Mail:
            if not sg_sender:
                raise HTTPException(status_code=500, detail="SENDGRID_SENDER_EMAIL is required in environment variables when using SendGrid.")
                
            message = Mail(
                from_email=sg_sender,
                to_emails=to_email,
                subject=f"GeoTagging Log Export — {len(records_list)} records",
                html_content=html
            )
            sg = SendGridAPIClient(sg_api_key)
            sg.send(message)
            return {"ok": True, "message": f"Log exported to {to_email} via SendGrid", "count": len(records_list)}
            
        # Fallback to standard SMTP (e.g. Ethereal Testing)
        else:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = f"GeoTagging Log Export — {len(records_list)} records"
            msg["From"]    = smtp_user
            msg["To"]      = to_email
            msg.attach(MIMEText(html, "html"))

            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, to_email, msg.as_string())

            return {"ok": True, "message": f"Log exported to {to_email} via SMTP", "count": len(records_list)}
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email send failed: {str(e)}")


# ════════════════════════════════════════════════════════════════════════════
#  ENHANCEMENT FEATURES (F1–F11)
# ════════════════════════════════════════════════════════════════════════════

# Service-level / scheduling metadata
_SERVICE_LEVELS = ("asap", "next_day", "standard", "special")
_PRIORITY = {"asap": 0, "next_day": 1, "special": 2, "standard": 3}
_NEARBY_DEFAULT_FT = 200.0


def _group_key(ph):
    return ph.location_group_id or ph.id


def _master_pin_dict(group_photos):
    """Collapse a list of attempt-Photos sharing a group into one master pin."""
    attempts = sorted(group_photos, key=lambda p: p.id)
    root = attempts[0]
    statuses = {p.status or "in_progress" for p in attempts}
    if statuses == {"archived"}:
        status = "archived"
    elif "in_progress" in statuses or "open" in statuses:
        status = "in_progress"
    else:
        status = "completed"
    pay = next((p.pay_rate for p in attempts if p.pay_rate is not None), None)

    # Service level: use the category of the most recent *active* attempt so
    # that adding a new attempt (e.g. "standard") to an existing ASAP group
    # reflects the latest intent rather than the original first attempt.
    active = [p for p in attempts if (p.status or "in_progress") not in ("archived", "completed")]
    service_level = (active[-1] if active else attempts[-1]).category or "standard"

    return {
        "location_group_id": _group_key(root),
        "latitude":          root.latitude,
        "longitude":         root.longitude,
        "address":           root.address,
        "zip_code":          root.zip_code,
        "service_level":     service_level,
        "pay_rate":          pay,
        "status":            status,
        "attempt_count":     len(attempts),
        "first_attempt_at":  _to_pst_iso(root.original_timestamp or root.timestamp),
        "last_attempt_at":   _to_pst_iso(attempts[-1].timestamp),
        "attempts":          [_photo_dict(p) for p in attempts],
    }


def _all_master_pins(db):
    photos = db.query(Photo).all()
    groups = {}
    for ph in photos:
        groups.setdefault(_group_key(ph), []).append(ph)
    return [_master_pin_dict(g) for g in groups.values()]


# ─── F1: DUPLICATE DETECTION / EXISTING PIN REUSE ───────────────────────────

@app.get("/locations/nearby")
def nearby_locations(lat: float, lng: float, radius_ft: float = _NEARBY_DEFAULT_FT):
    """Return existing master pins within `radius_ft` feet of (lat, lng)."""
    db = SessionLocal()
    pins = _all_master_pins(db)
    db.close()
    out = []
    for pin in pins:
        if pin["latitude"] is None or pin["longitude"] is None:
            continue
        dist = _haversine_ft(lat, lng, pin["latitude"], pin["longitude"])
        if dist <= radius_ft:
            pin["distance_ft"] = round(dist, 1)
            out.append(pin)
    out.sort(key=lambda p: p["distance_ft"])
    return out


@app.get("/locations")
def list_locations(service_levels: Optional[str] = None, status: Optional[str] = None):
    """Master-pin list with optional service-level (comma list) + status filters."""
    db = SessionLocal()
    pins = _all_master_pins(db)
    db.close()
    if service_levels:
        wanted = {s.strip().lower() for s in service_levels.split(",") if s.strip()}
        pins = [p for p in pins if p["service_level"] in wanted]
    if status:
        pins = [p for p in pins if p["status"] == status]
    return pins


@app.get("/locations/{group_id}/attempts")
def location_attempts(group_id: int):
    """Full attempt history / timeline for one master pin."""
    db = SessionLocal()
    photos = db.query(Photo).filter(
        (Photo.location_group_id == group_id) | (Photo.id == group_id)
    ).all()
    if not photos:
        db.close()
        raise HTTPException(status_code=404, detail="Location not found")
    result = _master_pin_dict(photos)   # build while session is open (lazy profiles)
    db.close()
    return result


# ─── F4: SERVICE-LEVEL SCHEDULING QUEUES ────────────────────────────────────

@app.get("/schedule")
def get_schedule(
    queue: Optional[str] = None,
    user_lat: Optional[float] = None,
    user_lng: Optional[float] = None,
):
    """Scheduling queues derived from service level.
    When user_lat/user_lng are provided each pin gets a distance_mi field and
    results are sorted nearest-first within each priority tier."""
    db = SessionLocal()
    pins = [p for p in _all_master_pins(db) if p["status"] != "archived"]
    db.close()
    if queue:
        q = queue.strip().lower()
        pins = [p for p in pins if p["service_level"] == q]
    for p in pins:
        p["priority"] = _PRIORITY.get(p["service_level"], 3)
        if user_lat is not None and user_lng is not None and p["latitude"] and p["longitude"]:
            dist_ft = _haversine_ft(user_lat, user_lng, p["latitude"], p["longitude"])
            p["distance_mi"] = round(dist_ft / 5280, 2)
        else:
            p["distance_mi"] = None
    # Sort: priority tier first, then nearest first (None distances go last)
    pins.sort(key=lambda p: (
        p["priority"],
        p["distance_mi"] if p["distance_mi"] is not None else float("inf"),
    ))
    # Group into named queues preserving the sorted order
    queues = {lvl: [] for lvl in _SERVICE_LEVELS}
    for p in pins:
        queues.setdefault(p["service_level"], []).append(p)
    return {"queues": queues, "items": pins}


# ─── F6: TIMESTAMP EDIT (10-MIN WINDOW) + AUDIT ─────────────────────────────

_EDIT_WINDOW_MIN = 10


@app.patch("/photos/{photo_id}/timestamp")
async def edit_timestamp(photo_id: int, data: dict = Body(...)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    # Window is measured from PIN-CREATION time (not capture time): the user
    # has 10 min after creating the pin to correct the timestamp.
    anchor = photo.created_at or photo.original_timestamp or photo.timestamp or datetime.utcnow()
    age_min = (datetime.utcnow() - anchor).total_seconds() / 60
    if age_min > _EDIT_WINDOW_MIN:
        db.close()
        raise HTTPException(status_code=423,
                            detail=f"Timestamp locked — edit window of {_EDIT_WINDOW_MIN} min elapsed")
    new_ts = _parse_device_ts(data.get("timestamp"))
    if not new_ts:
        db.close()
        raise HTTPException(status_code=422, detail="Valid 'timestamp' (ISO) required")
    old = photo.timestamp
    photo.timestamp = new_ts
    photo.edited_timestamp = new_ts
    db.add(TimestampEdit(photo_id=photo.id, old_value=old, new_value=new_ts,
                         edited_by=data.get("user_id")))
    db.commit()
    db.close()
    return {"ok": True, "timestamp": _to_pst_iso(new_ts)}


@app.get("/photos/{photo_id}/timestamp-history")
def timestamp_history(photo_id: int):
    db = SessionLocal()
    edits = db.query(TimestampEdit).filter(TimestampEdit.photo_id == photo_id).order_by(
        TimestampEdit.edited_at).all()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    result = {
        "photo_id":           photo_id,
        "original_timestamp": _to_pst_iso(photo.original_timestamp) if photo else None,
        "current_timestamp":  _to_pst_iso(photo.timestamp) if photo else None,
        "edits": [{
            "old_value": _to_pst_iso(e.old_value),
            "new_value": _to_pst_iso(e.new_value),
            "edited_by": e.edited_by,
            "edited_at": _to_pst_iso(e.edited_at),
        } for e in edits],
    }
    db.close()
    return result


# ─── F7: PAY RATE ───────────────────────────────────────────────────────────

@app.patch("/photos/{photo_id}/pay-rate")
async def update_pay_rate(photo_id: int, data: dict = Body(...)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    try:
        photo.pay_rate = int(round(float(data.get("pay_rate"))))
    except (TypeError, ValueError):
        db.close()
        raise HTTPException(status_code=422, detail="pay_rate must be a whole dollar number")
    # Pay rate is recorded but does NOT auto-complete the job.
    # Status is only changed by an explicit PATCH /photos/{id}/status call.
    db.commit()
    pay_val = photo.pay_rate
    db.close()
    return {"ok": True, "pay_rate": pay_val, "status": photo.status}


# ─── F10: ARCHIVE / JOB STATUS WORKFLOW ─────────────────────────────────────

@app.patch("/photos/{photo_id}/status")
async def update_status(photo_id: int, data: dict = Body(...)):
    db = SessionLocal()
    photo = db.query(Photo).filter(Photo.id == photo_id).first()
    if not photo:
        db.close()
        raise HTTPException(status_code=404, detail="Photo not found")
    st = (data.get("status") or "").strip().lower()
    if st not in ("open", "in_progress", "completed", "archived"):
        db.close()
        raise HTTPException(status_code=422, detail="status must be open|in_progress|completed|archived")
    old_status = photo.status or "open"
    # Any transition is allowed (forward, backward, or jump) — the client
    # confirms the change and can revert by sending the previous status back.
    photo.status = st
    if st in ("completed", "archived") and not photo.completed_at:
        photo.completed_at = datetime.utcnow()
    # Moving back out of a terminal state clears the completion stamp so
    # earnings/archive views stay consistent with the live status.
    if st in ("open", "in_progress"):
        photo.completed_at = None
    # Record the transition for the audit trail / undo.
    db.add(StatusHistory(
        photo_id=photo_id,
        old_status=old_status,
        new_status=st,
        updated_by=data.get("updated_by"),
    ))
    db.commit()
    db.close()
    return {"ok": True, "status": st, "previous_status": old_status}


@app.patch("/attempts/{attempt_id}/status")
async def update_attempt_status(attempt_id: int, data: dict = Body(...)):
    """Cascades a status change to the Attempt row and every sibling Photo
    row in one transaction, mirroring update_status's validation/transition
    logic (~L1726) so the audit trail and completion stamps stay consistent
    whether a job is closed out photo-by-photo or all at once."""
    db = SessionLocal()
    attempt = db.query(Attempt).filter(Attempt.id == attempt_id).first()
    if not attempt:
        db.close()
        raise HTTPException(status_code=404, detail="Attempt not found")
    st = (data.get("status") or "").strip().lower()
    if st not in ("open", "in_progress", "completed", "archived"):
        db.close()
        raise HTTPException(status_code=422, detail="status must be open|in_progress|completed|archived")

    # Update the Attempt row itself.
    attempt.status = st
    if st in ("completed", "archived") and not attempt.completed_at:
        attempt.completed_at = datetime.utcnow()
    if st in ("open", "in_progress"):
        attempt.completed_at = None

    # Cascade the same transition to every sibling Photo row.
    photos = db.query(Photo).filter(Photo.attempt_id == attempt_id).all()
    for photo in photos:
        old_status = photo.status or "open"
        photo.status = st
        if st in ("completed", "archived") and not photo.completed_at:
            photo.completed_at = datetime.utcnow()
        if st in ("open", "in_progress"):
            photo.completed_at = None
        db.add(StatusHistory(
            photo_id=photo.id,
            old_status=old_status,
            new_status=st,
            updated_by=data.get("updated_by"),
        ))
    db.commit()
    db.close()
    return {"ok": True, "status": st, "photos_updated": len(photos)}


@app.get("/photos/{photo_id}/status-history")
def status_history(photo_id: int):
    db = SessionLocal()
    rows = (
        db.query(StatusHistory)
        .filter(StatusHistory.photo_id == photo_id)
        .order_by(StatusHistory.updated_at.desc())
        .all()
    )
    db.close()
    return [
        {
            "id": r.id,
            "old_status": r.old_status,
            "new_status": r.new_status,
            "updated_by": r.updated_by,
            "updated_at": _to_pst_iso(r.updated_at),
        }
        for r in rows
    ]


@app.get("/archive")
def get_archive(search: Optional[str] = None, service_level: Optional[str] = None,
                status: str = "archived"):
    """Job list for the Archive screen, filtered by lifecycle status.

    `status` ∈ active (open + in_progress) | open | in_progress | completed | archived
    (defaults to "archived" for backward compatibility)."""
    db = SessionLocal()
    st = (status or "archived").strip().lower()
    if st == "active":
        status_filter = Photo.status.in_(("open", "in_progress"))
    elif st in ("open", "in_progress", "completed", "archived"):
        status_filter = Photo.status == st
    else:
        status_filter = Photo.status == "archived"
    photos = db.query(Photo).filter(status_filter).order_by(
        Photo.timestamp.desc()).all()
    result = []
    for ph in photos:
        d = _photo_dict(ph)   # build while session is open (lazy profiles)
        if service_level and d["category"] != service_level:
            continue
        if search:
            s = search.lower()
            hay = " ".join(str(x or "") for x in
                           [ph.note, ph.address, ph.zip_code,
                            *[p["name"] for p in d["profiles"]]]).lower()
            if s not in hay:
                continue
        result.append(d)
    db.close()
    return result


# ─── F8 / F9: EARNINGS + PAYOUTS / TIMESHEETS ───────────────────────────────

def _period_bounds(period):
    """Return (start_utc, label) for today|week|biweekly|month."""
    now = datetime.now(PST)
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    if period == "today":
        start = today
    elif period == "week":
        start = today - timedelta(days=today.weekday())
    elif period == "biweekly":
        start = today - timedelta(days=14)
    elif period == "month":
        start = today.replace(day=1)
    else:
        start = today
    return start.astimezone(timezone.utc).replace(tzinfo=None)


def _dedupe_by_attempt(jobs):
    """One Attempt can have several Photo rows (each upload = one row); dedupe
    so earnings/payouts count an attempt's pay_rate once, not once per photo."""
    seen = set()
    out = []
    for j in jobs:
        key = j.attempt_id or f"photo:{j.id}"
        if key in seen:
            continue
        seen.add(key)
        out.append(j)
    return out


@app.get("/earnings/summary")
def earnings_summary(period: str = "today", user_id: Optional[int] = None,
                      start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Aggregate completed-job earnings for the period (Uber-style).

    Counts jobs with status completed OR archived (an archived job that was
    completed must keep counting toward earnings/payouts). Pass `start_date`
    / `end_date` (YYYY-MM-DD) for a custom range — overrides `period`."""
    db = SessionLocal()
    q = db.query(Photo).filter(Photo.status.in_(("completed", "archived")),
                               Photo.completed_at.isnot(None))
    if start_date or end_date:
        period = "custom"
        start = _parse_device_ts(start_date)
        if start:
            q = q.filter(Photo.completed_at >= start)
        end = _parse_device_ts(end_date)
        if end:
            end = end.replace(hour=23, minute=59, second=59, microsecond=999999)
            q = q.filter(Photo.completed_at <= end)
    else:
        start = _period_bounds(period)
        q = q.filter(Photo.completed_at >= start)
    if user_id is not None:
        q = q.filter(Photo.user_id == user_id)
    jobs = q.all()
    jobs = _dedupe_by_attempt(jobs)
    total = sum((j.pay_rate or 0) for j in jobs)
    count = len(jobs)
    # daily breakdown for trend chart
    daily = {}
    for j in jobs:
        day = (_to_pst_iso(j.completed_at) or "")[:10]
        daily[day] = daily.get(day, 0) + (j.pay_rate or 0)

    def _job_summary(j):
        if j is None:
            return None
        return {
            "id":           j.id,
            "pay_rate":     j.pay_rate or 0,
            "profile_name": (j.profiles[0].name if j.profiles else "Unknown"),
            "category":     j.category or "standard",
            "address":      j.address,
            "completed_at": _to_pst_iso(j.completed_at),
        }

    highest = _job_summary(max(jobs, key=lambda j: (j.pay_rate or 0), default=None))
    lowest  = _job_summary(min(jobs, key=lambda j: (j.pay_rate or 0), default=None))

    # Pay still assigned to unfinished profiles — not a cash-out / "available
    # to withdraw" balance, and not scoped to the selected period.
    open_jobs = []
    for p in db.query(Profile).all():
        if not p.pay_rate:
            continue
        attempts = list(getattr(p, "attempts", None) or [])
        if attempts:
            if all((a.status or "") in ("completed", "archived") for a in attempts):
                continue
        else:
            photos = list(p.photos or [])
            if photos and all(
                (ph.status or "") in ("completed", "archived") for ph in photos
            ):
                continue
        open_jobs.append({
            "id": p.id,
            "name": p.name,
            "pay_rate": p.pay_rate,
        })

    db.close()
    return {
        "period":             period,
        "jobs_completed":     count,
        "total_earnings":     total,
        "available_earnings": sum(j["pay_rate"] for j in open_jobs),
        "open_jobs":          open_jobs,
        "average_per_job":    round(total / count, 2) if count else 0,
        "highest_paying_job": highest,
        "lowest_paying_job":  lowest,
        "daily_totals":       [{"date": d, "amount": a} for d, a in sorted(daily.items())],
    }


@app.get("/payouts")
def get_payouts(user_id: Optional[int] = None):
    """Closed-out pins grouped by completion day — works like the daily log but
    for completed jobs (Don #8). Each day lists its individual closed pins,
    its daily total, and a cumulative running total across all days.

    Includes archived jobs that were completed (archiving must not remove a
    job from its payout history)."""
    db = SessionLocal()
    q = db.query(Photo).filter(Photo.status.in_(("completed", "archived")),
                               Photo.completed_at.isnot(None))
    if user_id is not None:
        q = q.filter(Photo.user_id == user_id)
    jobs = q.order_by(Photo.completed_at.desc()).all()
    jobs = _dedupe_by_attempt(jobs)
    # Build per-day buckets with the actual closed-pin entries (like the log)
    days = {}
    for j in jobs:
        day = (_to_pst_iso(j.completed_at) or "")[:10]
        bucket = days.setdefault(day, {"date": day, "jobs": 0, "amount": 0, "entries": []})
        bucket["jobs"]   += 1
        bucket["amount"] += (j.pay_rate or 0)
        bucket["entries"].append({
            "id":           j.id,
            "profile_name": (j.profiles[0].name if j.profiles else "Unknown"),
            "category":     j.category or "standard",
            "pay_rate":     j.pay_rate or 0,
            "address":      j.address,
            "image_url":    j.image_url,
            "completed_at": _to_pst_iso(j.completed_at),
        })
    db.close()
    # Newest day first; keep a cumulative running total (oldest→newest), Uber-style
    daily = sorted(days.values(), key=lambda d: d["date"], reverse=True)
    cumulative = 0
    for d in sorted(daily, key=lambda d: d["date"]):     # ascending for running sum
        cumulative += d["amount"]
        d["running_total"] = cumulative
    return {
        "total_earnings": sum(d["amount"] for d in daily),
        "total_jobs":     sum(d["jobs"] for d in daily),
        "daily":          daily,
    }


@app.post("/payouts/finalize")
async def finalize_payout(data: dict = Body(...)):
    """Snapshot a pay period so historical totals don't drift."""
    db = SessionLocal()
    start = _parse_device_ts(data.get("period_start"))
    end = _parse_device_ts(data.get("period_end"))
    uid = data.get("user_id")
    q = db.query(Photo).filter(Photo.status.in_(("completed", "archived")),
                               Photo.completed_at.isnot(None))
    if uid is not None:
        q = q.filter(Photo.user_id == uid)
    if start:
        q = q.filter(Photo.completed_at >= start)
    if end:
        q = q.filter(Photo.completed_at <= end)
    jobs = q.all()
    jobs = _dedupe_by_attempt(jobs)
    snap = PayoutPeriod(
        user_id=uid, period_start=start, period_end=end,
        total_amount=sum((j.pay_rate or 0) for j in jobs), jobs_count=len(jobs))
    db.add(snap)
    db.commit()
    db.refresh(snap)
    out = {"id": snap.id, "total_amount": snap.total_amount, "jobs_count": snap.jobs_count}
    db.close()
    return out


# ─── F5: DRAFT AUTO-SAVE ────────────────────────────────────────────────────

@app.put("/drafts")
async def save_draft(data: dict = Body(...)):
    """Upsert a pin draft (continuous autosave from mobile)."""
    draft_id = (data.get("id") or str(uuid.uuid4()))
    db = SessionLocal()
    draft = db.query(PinDraft).filter(PinDraft.id == draft_id).first()
    if not draft:
        draft = PinDraft(id=draft_id)
        db.add(draft)
    draft.user_id    = data.get("user_id")
    draft.payload    = json.dumps(data.get("payload", {}))
    draft.updated_at = datetime.utcnow()
    db.commit()
    db.close()
    return {"ok": True, "id": draft_id}


@app.get("/drafts")
def list_drafts(user_id: Optional[int] = None):
    db = SessionLocal()
    q = db.query(PinDraft)
    if user_id is not None:
        q = q.filter(PinDraft.user_id == user_id)
    drafts = q.order_by(PinDraft.updated_at.desc()).all()
    out = [{"id": d.id, "user_id": d.user_id,
            "payload": json.loads(d.payload or "{}"),
            "updated_at": _to_pst_iso(d.updated_at)} for d in drafts]
    db.close()
    return out


@app.delete("/drafts/{draft_id}")
def delete_draft(draft_id: str):
    db = SessionLocal()
    draft = db.query(PinDraft).filter(PinDraft.id == draft_id).first()
    if draft:
        db.delete(draft)
        db.commit()
    db.close()
    return {"ok": True}


# ─── F11: SAVED EMAIL RECIPIENTS ────────────────────────────────────────────

@app.get("/recipients")
def list_recipients(user_id: Optional[int] = None):
    db = SessionLocal()
    q = db.query(EmailRecipient)
    if user_id is not None:
        q = q.filter(EmailRecipient.user_id == user_id)
    rows = q.all()
    out = [{"id": r.id, "label": r.label, "email": r.email, "user_id": r.user_id} for r in rows]
    db.close()
    return out


@app.post("/recipients")
async def add_recipient(data: dict = Body(...)):
    email = (data.get("email") or "").strip()
    if "@" not in email:
        raise HTTPException(status_code=422, detail="Valid email required")
    db = SessionLocal()
    # cap at 10 per user
    uid = data.get("user_id")
    count = db.query(EmailRecipient).filter(EmailRecipient.user_id == uid).count()
    if count >= 10:
        db.close()
        raise HTTPException(status_code=422, detail="Maximum of 10 saved recipients reached")
    r = EmailRecipient(user_id=uid, label=(data.get("label") or "").strip() or None, email=email)
    db.add(r)
    db.commit()
    db.refresh(r)
    out = {"id": r.id, "label": r.label, "email": r.email}
    db.close()
    return out


@app.patch("/recipients/{rid}")
async def edit_recipient(rid: int, data: dict = Body(...)):
    db = SessionLocal()
    r = db.query(EmailRecipient).filter(EmailRecipient.id == rid).first()
    if not r:
        db.close()
        raise HTTPException(status_code=404, detail="Recipient not found")
    if "email" in data:
        r.email = data["email"].strip()
    if "label" in data:
        r.label = (data["label"] or "").strip() or None
    db.commit()
    out = {"id": r.id, "label": r.label, "email": r.email}
    db.close()
    return out


@app.delete("/recipients/{rid}")
def delete_recipient(rid: int):
    db = SessionLocal()
    r = db.query(EmailRecipient).filter(EmailRecipient.id == rid).first()
    if r:
        db.delete(r)
        db.commit()
    db.close()
    return {"ok": True}


# ─── F11: EXCEL EXPORT (replaces multi-format exports) ──────────────────────

# Multi-job export (selecting & sending many jobs to one recipient) — the
# canonical 5-column log layout the client signed off on.
_EXPORT_COLUMNS = [
    ("id_ctrl",         "ID & Cntrl #"),
    ("date_time",       "Date & Time"),
    ("service_ordered", "Service Ordered"),
    ("address",         "Address"),
    ("coordinates",     "Lat / Long"),
    ("detailed_notes",  "Detailed Notes"),
]

# Single-job export — same shape as above plus the few extras the client asked
# for on individual records (coordinates, status, agent). The watermarked photo
# rides along as a separate attachment.
_EXPORT_COLUMNS_JOB = [
    ("id_ctrl",         "ID & Cntrl #"),
    ("date_time",       "Date & Time"),
    ("service_ordered", "Service Ordered"),
    ("address",         "Address"),
    ("coordinates",     "Lat / Long"),
    ("job_status",      "Job Status"),
    ("agent",           "Agent"),
    ("detailed_notes",  "Detailed Notes"),
]

_COL_WIDTHS = {
    "ID & Cntrl #": 24, "Date & Time": 22, "Service Ordered": 18,
    "Address": 32, "Lat / Long": 24, "Job Status": 16, "Agent": 18,
    "Detailed Notes": 36,
    # Profile-list export (manual selection vs full list)
    "File Number": 14, "Name": 22, "Priority Level": 16, "Notes": 36,
    "Photo": 16,
}

# Profile-list export — the full-list (unselected) variant. One row per profile,
# built from that profile's most recent photo. No embedded photo.
_EXPORT_COLUMNS_LIST = [
    ("file_number",    "File Number"),
    ("name",           "Name"),
    ("date_time",      "Date & Time"),
    ("priority_level", "Priority Level"),
    ("address",        "Address"),
    ("notes",          "Notes"),
]

# Manual-selection variant — same fields plus the most recent photo for each
# profile, watermarked with its timestamp + geotag and embedded into the sheet.
_EXPORT_COLUMNS_MANUAL = _EXPORT_COLUMNS_LIST + [("photo", "Photo")]

# Map a stored priority category to the label shown in exports.
_PRIORITY_LABELS = {
    "asap": "ASAP", "rush": "ASAP", "next_day": "Next Day",
    "standard": "Standard", "special": "Special", "airport": "Airport",
}


def _priority_label(category):
    return _PRIORITY_LABELS.get((category or "standard").lower(), "Standard")


def _build_xlsx(records, columns=_EXPORT_COLUMNS, sheet_title="Activity Log",
                base_name="log_export"):
    """Return (bytes, filename) for an .xlsx (or CSV fallback) of the records,
    using the supplied (key, header) column set."""
    if Workbook is None:
        # Fallback: CSV bytes
        buf = io.StringIO()
        w = _csv.writer(buf)
        w.writerow([h for _, h in columns])
        for r in records:
            w.writerow([r.get(k, "") for k, _ in columns])
        return buf.getvalue().encode(), f"{base_name}.csv"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    for col, (_, label) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
    for ri, r in enumerate(records, start=2):
        for ci, (key, _) in enumerate(columns, start=1):
            ws.cell(row=ri, column=ci, value=r.get(key, ""))
    for col, (_, label) in enumerate(columns, start=1):
        ws.column_dimensions[chr(64 + col)].width = _COL_WIDTHS.get(label, 20)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), f"{base_name}.xlsx"


# Bundled first (see fonts/NOTICE.md) so the watermark renders legibly on any
# server regardless of which system fonts happen to be installed — the system
# paths below only matter as a fallback if the bundled file is ever missing.
# Bold reads far more clearly than regular weight against busy photo
# backgrounds, which is why it's listed ahead of the regular-weight fallbacks.
_FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fonts")
_FONT_PATHS = [
    os.path.join(_FONT_DIR, "DejaVuSans-Bold.ttf"),
    os.path.join(_FONT_DIR, "DejaVuSans.ttf"),
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
    "/Library/Fonts/Arial Bold.ttf",
    "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
    "DejaVuSans-Bold.ttf",
]


def _load_font(size):
    from PIL import ImageFont
    for p in _FONT_PATHS:
        try:
            return ImageFont.truetype(p, size)
        except Exception:
            continue
    # Only reached if every bundled + system path above failed to load.
    # `size=` keeps it legible instead of Pillow's tiny fixed-size bitmap font.
    return ImageFont.load_default(size=size)


def _watermark_photo_png(image_path, caption_lines, max_w=300, max_h=400):
    """Return (png_bytes, width_px, height_px): the photo scaled to fit
    (max_w × max_h) with `caption_lines` (timestamp + geotag) burned into a
    translucent bar along the bottom. Font size scales with the output width so
    the caption is legible on both thumbnail (xlsx) and full-size (service
    record) exports. Returns None if Pillow is unavailable or the file can't be
    read. Doing this server-side (Pillow) is reliable for any JPEG/PNG,
    regardless of the client device's image codec."""
    try:
        from PIL import Image as PILImage, ImageDraw, ImageOps, ImageFile
        # Phone photos (e.g. Samsung) sometimes arrive a few bytes short; without
        # this Pillow raises "image file is truncated" and the photo drops out.
        ImageFile.LOAD_TRUNCATED_IMAGES = True
    except Exception:
        return None
    try:
        img = PILImage.open(image_path)
        img = ImageOps.exif_transpose(img).convert("RGB")
    except Exception:
        return None
    # Scale to fit the target box, preserving aspect ratio (never upscale).
    ratio = min(max_w / img.width, max_h / img.height, 1.0)
    if ratio < 1.0:
        img = img.resize((max(1, int(img.width * ratio)),
                          max(1, int(img.height * ratio))))
    draw = ImageDraw.Draw(img, "RGBA")
    fsize = max(12, min(40, img.width // 34))   # ~legible at any width
    font = _load_font(fsize)
    lines = [ln for ln in caption_lines if ln]
    if lines:
        pad = max(6, fsize // 2)
        line_h = fsize + max(4, fsize // 3)
        box_h = pad * 2 + line_h * len(lines)
        draw.rectangle([0, img.height - box_h, img.width, img.height],
                       fill=(0, 0, 0, 160))
        y = img.height - box_h + pad
        for ln in lines:
            bbox = draw.textbbox((0, 0), ln, font=font)
            tw = bbox[2] - bbox[0]
            x = max(0, (img.width - tw) / 2)
            # 2px shadow keeps white text legible on bright photos.
            draw.text((x + 2, y + 2), ln, fill=(0, 0, 0, 200), font=font)
            draw.text((x, y), ln, fill=(255, 255, 255, 255), font=font)
            y += line_h
    out = io.BytesIO()
    img.save(out, format="PNG")
    return out.getvalue(), img.width, img.height


def _street_zip(address, zip_code=None):
    """Full location for watermarks and email: street, city, state, ZIP.
    Does not drop city/state after the first comma."""
    full = re.sub(r"\s+", " ", (address or "").strip())
    zip_val = (zip_code or "").strip()
    if not zip_val:
        m = re.search(r"\b\d{5}(?:-\d{4})?\b", full)
        zip_val = m.group(0) if m else ""
    if not full:
        return zip_val
    if zip_val and not re.search(rf"\b{re.escape(zip_val)}\b", full):
        return f"{full}, {zip_val}"
    return full


def _is_file_number_na(value):
    """True when file number is missing or the explicit N/A sentinel."""
    v = (value or "").strip()
    return not v or v.upper() == "N/A"


def _photo_caption(ph):
    """Caption lines burned into an exported photo:
      FILE-123                 (file number, or profile name when N/A / empty)
      2025-07-03 14:30 PST
      4822 Reno Drive, San Diego, CA 92101
      32.690861, -117.113289
    Priority / service level is intentionally omitted."""
    lines = []
    file_no = (ph.file_number or "").strip()
    if not _is_file_number_na(file_no):
        lines.append(file_no)
    else:
        name = ""
        try:
            if ph.profiles:
                name = (ph.profiles[0].name or "").strip()
        except Exception:
            name = ""
        if not name and ph.profile_id:
            try:
                db = SessionLocal()
                prof = db.query(Profile).filter(Profile.id == ph.profile_id).first()
                name = (prof.name or "").strip() if prof else ""
                db.close()
            except Exception:
                name = ""
        if name:
            lines.append(name)
    ts = _to_pst_iso(ph.taken_at or ph.timestamp)
    if ts:
        # 2025-07-03T14:30:45-07:00 → "2025-07-03 14:30 PST"
        lines.append(ts.replace("T", " ")[:16] + " PST")
    street_zip = _street_zip(ph.address, ph.zip_code)
    if street_zip:
        lines.append(street_zip)
    if ph.latitude is not None and ph.longitude is not None:
        if abs(ph.latitude) > 0.0001 or abs(ph.longitude) > 0.0001:
            lines.append(f"{ph.latitude:.6f}, {ph.longitude:.6f}")
    return lines


def _watermarked_photo_payloads(photo_ids, max_w=1600, max_h=1600):
    """[{filename, content_b64, mimetype}] for each photo id, timestamp +
    geotag already burned in. Used as regular email attachments and as the
    share-sheet payload so the spreadsheet is no longer required."""
    payloads = []
    for pid in photo_ids or []:
        img_bytes = _get_watermarked_bytes(pid, max_w=max_w, max_h=max_h)
        if not img_bytes:
            continue
        payloads.append({
            "filename": f"photo_{pid}.png",
            "content_b64": base64.b64encode(img_bytes).decode(),
            "mimetype": "image/png",
        })
    return payloads


def _get_watermarked_bytes(photo_id, max_w=800, max_h=800):
    """Fetch + watermark a photo's bytes directly (no HTTP round-trip) so it
    can be attached inline (CID) in an export email — this travels with the
    message itself, so it still renders even when a mail client refuses to
    fetch remote images (e.g. for an unauthenticated/spam-flagged sender)."""
    db = SessionLocal()
    ph = db.query(Photo).filter(Photo.id == photo_id).first()
    if not ph or not ph.image_url:
        db.close()
        return None
    # Build caption while the session is open so profile names can lazy-load.
    caption = _photo_caption(ph)
    path = ph.image_url.lstrip("/")
    db.close()
    if not os.path.exists(path):
        return None
    stamped = _watermark_photo_png(path, caption, max_w=max_w, max_h=max_h)
    return stamped[0] if stamped else None


def _sendgrid_inline_attachments(photo_ids, max_w=800, max_h=800):
    """SendGrid Attachment objects (inline, Content-ID) for the given photo
    ids — matches the `cid:photo_{id}` references built into the HTML body."""
    from sendgrid.helpers.mail import (
        Attachment, FileContent, FileName, FileType, Disposition, ContentId)
    atts = []
    for pid in photo_ids:
        img_bytes = _get_watermarked_bytes(pid, max_w=max_w, max_h=max_h)
        if not img_bytes:
            continue
        atts.append(Attachment(
            FileContent(base64.b64encode(img_bytes).decode()),
            FileName(f"photo_{pid}.png"),
            FileType("image/png"),
            Disposition("inline"),
            ContentId(f"photo_{pid}"),
        ))
    return atts


def _smtp_inline_photo_parts(photo_ids, max_w=800, max_h=800):
    """MIMEImage parts (inline, Content-ID) for the given photo ids — matches
    the `cid:photo_{id}` references built into the HTML body."""
    from email.mime.image import MIMEImage
    parts = []
    for pid in photo_ids:
        img_bytes = _get_watermarked_bytes(pid, max_w=max_w, max_h=max_h)
        if not img_bytes:
            continue
        img = MIMEImage(img_bytes, _subtype="png")
        img.add_header("Content-ID", f"<photo_{pid}>")
        img.add_header("Content-Disposition", "inline", filename=f"photo_{pid}.png")
        parts.append(img)
    return parts


def _photo_link_cell(ws, row, col, photo_id, wrap):
    """Write a 'View photo' hyperlink (to the server-watermarked image) into a
    cell. Linking instead of embedding keeps the rows a normal height — no big
    blank space in the text columns."""
    cell = ws.cell(row=row, column=col)
    if photo_id is not None:
        cell.value = "View photo"
        cell.hyperlink = f"{PUBLIC_BASE_URL}/api/photos/{photo_id}/watermarked"
        cell.font = Font(color="0563C1", underline="single")
    else:
        cell.value = "(no photo)"
    cell.alignment = wrap


def _build_xlsx_with_photos(records):
    """Build the manual-selection .xlsx: the profile-list columns plus a Photo
    column that LINKS to each profile's most recent (watermarked) photo. Rows
    stay a normal height — no embedded image blowing up the row. `records` carry
    a `photo_id`. Falls back to the plain builder if openpyxl is missing."""
    if Workbook is None:
        return _build_xlsx(records, _EXPORT_COLUMNS_LIST,
                           sheet_title="Profiles Export",
                           base_name="profiles_export")

    columns = _EXPORT_COLUMNS_MANUAL
    photo_col = len(columns)  # Photo is the last column (1-indexed)

    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles Export"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    for col, (_, label) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        ws.column_dimensions[chr(64 + col)].width = _COL_WIDTHS.get(label, 20)

    _wrap = Alignment(wrap_text=True, vertical="top")
    for ri, r in enumerate(records, start=2):
        for ci, (key, _) in enumerate(columns, start=1):
            if key == "photo":
                continue
            cell = ws.cell(row=ri, column=ci, value=r.get(key, ""))
            cell.alignment = _wrap
        _photo_link_cell(ws, ri, photo_col, r.get("photo_id"), _wrap)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), "profiles_export.xlsx"


_SENDER_NAME = os.environ.get("EMAIL_SENDER_NAME", "The GeoTagging CRM Team")


def _pro_email_html(intro, *, sender=None, table_html="", count=None,
                    closing=None):
    """Wrap an export email in a professional layout: greeting, intro, optional
    table, a courteous closing and a 'Best regards' signature (from), plus a
    branded footer. Keeps every export email consistent and presentable."""
    when = datetime.now(PST).strftime('%b %d, %Y %I:%M %p PST')
    sign = _esc(sender) if sender else _SENDER_NAME
    meta = f"{count} record(s) &middot; " if count is not None else ""
    closing = closing or ("Please find the file attached. If you have any "
                          "questions or need anything further, simply reply "
                          "to this email.")
    return f"""<html><body style="font-family:Arial,Helvetica,sans-serif;color:#0f172a;font-size:14px;line-height:1.6;max-width:660px;margin:0 auto;">
  <p style="margin:0 0 14px;">Hello,</p>
  <p style="margin:0 0 14px;">{intro}</p>
  {table_html}
  <p style="margin:16px 0 14px;">{closing}</p>
  <p style="margin:0;">Best regards,</p>
  <p style="margin:2px 0 0;font-weight:700;">{sign}</p>
  <hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0 8px;">
  <p style="margin:0;color:#94a3b8;font-size:11px;">{meta}Generated by GeoTagging CRM &middot; {when}</p>
</body></html>"""


def _pro_email_text(intro, *, sender=None, count=None, closing=None):
    """Plain-text sibling of :func:`_pro_email_html` for text/plain parts."""
    sign = sender or _SENDER_NAME
    meta = f"{count} record(s)\n" if count is not None else ""
    closing = closing or ("Please find the file attached. If you have any "
                          "questions or need anything further, simply reply "
                          "to this email.")
    return (f"Hello,\n\n{intro}\n\n{closing}\n\n"
            f"Best regards,\n{sign}\n\n"
            f"{meta}Generated by GeoTagging CRM")


@app.post("/export/excel")
async def export_excel(request: Request):
    """Email the selected profiles to one or more recipients. Each record's
    fields and watermarked photo go in the message body — no spreadsheet."""
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        data = await request.json()
        recipients = data.get("recipients") or ([data["to"]] if data.get("to") else [])
        records = data.get("records", [])
    else:
        form = await request.form()
        recipients = json.loads(form.get("recipients", "[]")) or (
            [form.get("to")] if form.get("to") else [])
        try:
            records = json.loads(form.get("records", "[]"))
        except Exception:
            records = []

    recipients = [e.strip() for e in recipients if e and "@" in e]
    if not recipients:
        raise HTTPException(status_code=422, detail="At least one recipient email required")

    smtp_host = os.environ.get("SMTP_HOST")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    sg_api_key = os.environ.get("SENDGRID_API_KEY")
    sg_sender = os.environ.get("SENDGRID_SENDER_EMAIL")

    # Logged fields + watermarked photos live in the email body. A spreadsheet
    # is only generated as a fallback when outbound email isn't configured
    # (web download). Recipients never receive an .xlsx — that was removed.
    photo_ids = [r["photo_id"] for r in records if r.get("photo_id") is not None]
    body_html = _build_records_export_html(records)
    body_text = None
    file_bytes, fname = _build_xlsx_with_photos(records)
    b64 = base64.b64encode(file_bytes).decode()
    if len(records) == 1:
        rec = records[0]
        subject = f"GeoTagging CRM — {rec.get('name') or 'Profile'} Export"
    else:
        subject = f"GeoTagging CRM — Profiles Export ({len(records)} records)"

    if not sg_api_key and (not smtp_host or not smtp_user):
        return {"ok": True, "message": "Email not configured — file returned",
                "filename": fname, "file_base64": b64, "count": len(records)}

    try:
        if sg_api_key and SendGridAPIClient and Mail:
            message = Mail(from_email=sg_sender, to_emails=recipients,
                           subject=subject, html_content=body_html)
            for att in _sendgrid_inline_attachments(photo_ids):
                message.add_attachment(att)
            SendGridAPIClient(sg_api_key).send(message)
            return {"ok": True, "message": f"Export sent to {', '.join(recipients)}",
                    "count": len(records)}
        else:
            msg = MIMEMultipart("mixed")
            msg["Subject"] = subject
            msg["From"] = smtp_user
            msg["To"] = ", ".join(recipients)
            related = MIMEMultipart("related")
            alt = MIMEMultipart("alternative")
            if body_text:
                alt.attach(MIMEText(body_text, "plain"))
            alt.attach(MIMEText(body_html, "html"))
            related.attach(alt)
            for part in _smtp_inline_photo_parts(photo_ids):
                related.attach(part)
            msg.attach(related)
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, recipients, msg.as_string())
            return {"ok": True, "message": f"Export sent to {', '.join(recipients)}",
                    "count": len(records)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


def _esc(s):
    """Minimal HTML escape for values dropped into the email body."""
    return (str(s or "")
            .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def _fmt_note_stamp(date_time):
    """'YYYY-MM-DD HH:MM:SS' -> 'hh:mm am  -  MM/DD/YYYY'. Falls back to raw."""
    raw = str(date_time or "").strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            dt = datetime.strptime(raw, fmt)
            return f"{dt.strftime('%I:%M %p').lower().lstrip('0')}  -  {dt.strftime('%m/%d/%Y')}"
        except (ValueError, TypeError):
            continue
    return raw


def _build_records_export_html(records):
    """Build the profiles-export email body: every selected record's File
    Number, Name, Date & Time, Priority Level, Address and Notes, with its
    watermarked photo embedded inline — one block per record, directly in
    the message body. No spreadsheet attachment."""
    records = records or []
    blocks = []
    for rec in records:
        rows = [
            ("File Number",    _esc(rec.get("file_number")) or "—"),
            ("Name",           _esc(rec.get("name")) or "—"),
            ("Date & Time",    _esc(rec.get("date_time")) or "—"),
            ("Priority Level", _esc(rec.get("priority_level")) or "—"),
            ("Address",        _esc(rec.get("address")) or "—"),
        ]
        if rec.get("coordinates"):
            rows.append(("Lat / Long", _esc(rec.get("coordinates"))))
        rows_html = "".join(
            f'<tr><td style="padding:2px 12px 2px 0;color:#64748b;white-space:nowrap;">'
            f'{label}:</td><td style="padding:2px 0;color:#0f172a;font-weight:600;">'
            f'{value}</td></tr>'
            for label, value in rows
        )
        notes = (rec.get("notes") or "").strip()
        notes_html = (_esc(notes) if notes
                      else '<span style="color:#94a3b8;">No notes recorded.</span>')

        photo_html = ""
        photo_id = rec.get("photo_id")
        if photo_id is not None:
            # cid: reference — the image travels WITH the email as an inline
            # attachment (see _attach_inline_photos), so it renders even when
            # a client refuses to fetch remote images from this sender.
            photo_html = (
                f'<img src="cid:photo_{photo_id}" alt="Service photo" '
                f'style="max-width:100%;border-radius:8px;display:block;margin-top:10px;">'
            )

        blocks.append(f"""\
<table style="border-collapse:collapse;font-size:14px;margin:0 0 8px;">{rows_html}</table>
<div style="margin:0 0 4px;color:#334155;font-weight:700;font-size:13px;">Notes</div>
<div style="margin:0 0 10px;color:#0f172a;line-height:1.5;">{notes_html}</div>
{photo_html}""")

    separator = '<hr style="border:none;border-top:1px solid #e2e8f0;margin:20px 0;">'
    body = separator.join(blocks) if blocks else '<p>No records to export.</p>'
    return (f'<div style="font-family:Arial,Helvetica,sans-serif;color:#0f172a;'
            f'max-width:640px;">{body}</div>')


def _build_service_record_html(records, header=None, latest_only=False,
                                photo_id=None):
    """Build the single-profile service-record email body: File Number, Name,
    Date & Time, Priority Level, Address, Notes, and the watermarked Photo
    embedded inline (this export is always for one profile/job at a time)."""
    records = records or []
    header = header or {}
    latest = records[0] if records else {}

    def hv(key, fallback_key=None, default="—"):
        v = header.get(key)
        if not v and fallback_key:
            v = latest.get(fallback_key)
        return _esc(v) if v else default

    rows = [
        ("File Number",    hv("file_number")),
        ("Name",           hv("name", "id_ctrl")),
        ("Date & Time",    hv("date_time", "date_time")),
        ("Priority Level", hv("priority_level", "service_ordered", "Standard")),
        ("Address",        hv("address", "address")),
        ("Lat / Long",     hv("coordinates", "coordinates")),
    ]
    header_html = "".join(
        f'<tr><td style="padding:2px 12px 2px 0;color:#64748b;white-space:nowrap;">'
        f'{label}:</td><td style="padding:2px 0;color:#0f172a;font-weight:600;">'
        f'{value}</td></tr>'
        for label, value in rows
    )

    # Notes — one attempt when latest_only, otherwise every attempt oldest-first.
    note_records = ([latest] if latest_only else list(reversed(records)))
    note_lines = []
    for r in note_records:
        note = (r.get("detailed_notes") or r.get("notes") or "").strip()
        if not note:
            continue
        stamp = _fmt_note_stamp(r.get("date_time"))
        prefix = f"{stamp}  -  " if stamp else ""
        note_lines.append(
            f'<div style="margin:0 0 10px;color:#0f172a;line-height:1.5;">'
            f'<span style="color:#475569;">{_esc(prefix)}</span>{_esc(note)}</div>'
        )
    notes_html = ("".join(note_lines)
                  or '<div style="color:#94a3b8;">No notes recorded.</div>')

    photo_html = ""
    if photo_id is not None:
        # cid: reference — see _attach_inline_photos for the matching attachment.
        photo_html = (
            f'<h3 style="margin:20px 0 10px;font-size:14px;letter-spacing:0.4px;'
            f'color:#334155;">Photo</h3>'
            f'<img src="cid:photo_{photo_id}" alt="Service photo" '
            f'style="max-width:100%;border-radius:8px;display:block;">'
        )

    return f"""\
<div style="font-family:Arial,Helvetica,sans-serif;color:#0f172a;max-width:640px;">
  <table style="border-collapse:collapse;font-size:14px;margin:0 0 16px;">{header_html}</table>
  <h3 style="margin:0 0 12px;font-size:14px;letter-spacing:0.4px;color:#334155;">Notes</h3>
  {notes_html}
  {photo_html}
</div>"""


def _job_photo_caption(rec):
    """Caption lines burned into a service-record photo, from the matching
    export record (already formatted by the client). Priority / service level
    is intentionally omitted from watermarks."""
    lines = []
    file_no = (rec.get("file_number") or "").strip()
    if not _is_file_number_na(file_no):
        heading = file_no
    else:
        heading = (rec.get("profile_name") or "").strip()
    if heading:
        lines.append(heading)
    if rec.get("date_time"):
        lines.append(str(rec["date_time"]))
    # Prefer an already street+ZIP-formatted address from the export record.
    if rec.get("address"):
        lines.append(str(rec["address"]))
    if rec.get("coordinates"):
        lines.append(str(rec["coordinates"]))
    if rec.get("agent"):
        lines.append(f"Agent: {rec['agent']}")
    return [ln for ln in lines if ln]


def _build_job_xlsx_with_photos(records, photo_ids, base_name="service-record"):
    """Service Record .xlsx with a trailing Photo column that LINKS to each
    attempt's (watermarked) photo. Linking keeps rows a normal height (no huge
    blank gap from an embedded image). `photo_ids` runs parallel to `records`."""
    if Workbook is None:
        return _build_xlsx(records, _EXPORT_COLUMNS_JOB,
                           sheet_title="Service Record", base_name=base_name)

    columns = _EXPORT_COLUMNS_JOB + [("photo", "Photo")]
    photo_col = len(columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Service Record"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    for col, (_, label) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        ws.column_dimensions[chr(64 + col)].width = _COL_WIDTHS.get(label, 20)

    _wrap = Alignment(wrap_text=True, vertical="top")
    for ri, r in enumerate(records, start=2):
        for ci, (key, _) in enumerate(columns, start=1):
            if key == "photo":
                continue
            cell = ws.cell(row=ri, column=ci, value=r.get(key, ""))
            cell.alignment = _wrap
        pid = photo_ids[ri - 2] if (ri - 2) < len(photo_ids) else None
        _photo_link_cell(ws, ri, photo_col, pid, _wrap)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), f"{base_name}.xlsx"


@app.post("/export/job")
async def export_job(request: Request):
    """Single-attempt export: email the logged fields plus a picture with
    timestamp and geotag burned in (inline in the body and attached as a
    PNG). Spreadsheets are not emailed — the client asked that removed.

    With no recipients (or when email isn't configured) watermarked photos
    are returned as base64 for the OS share sheet. An .xlsx is still
    generated in that fallback only, for the web download path.
    """
    data = await request.json()
    recipients = data.get("recipients") or ([data["to"]] if data.get("to") else [])
    recipients = [e.strip() for e in recipients if e and "@" in e]
    records = data.get("records", [])
    # attachments: [{"filename", "content_b64", "mimetype"}] — legacy client photos.
    attachments = data.get("attachments", [])
    photo_ids = data.get("photo_ids") or []
    subject = (data.get("subject") or "Service Record").strip()
    latest_only = bool(data.get("latest_only"))
    html_body = _build_service_record_html(
        records, data.get("header"), latest_only,
        photo_id=photo_ids[0] if photo_ids else None)

    photo_payloads = _watermarked_photo_payloads(photo_ids)
    if not photo_payloads:
        photo_payloads = [a for a in attachments if a.get("content_b64")]

    smtp_host = os.environ.get("SMTP_HOST")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    sg_api_key = os.environ.get("SENDGRID_API_KEY")
    sg_sender = os.environ.get("SENDGRID_SENDER_EMAIL")

    if not recipients or (not sg_api_key and (not smtp_host or not smtp_user)):
        return {"ok": True,
                "message": ("File generated" if not recipients
                            else "Email not configured — file returned"),
                "photos": photo_payloads, "count": len(records)}

    inline_photo_ids = [photo_ids[0]] if photo_ids else []
    try:
        if sg_api_key and SendGridAPIClient and Mail:
            from sendgrid.helpers.mail import (
                Attachment, FileContent, FileName, FileType, Disposition)
            atts = []
            for a in photo_payloads:
                atts.append(Attachment(
                    FileContent(a["content_b64"]),
                    FileName(a.get("filename", "photo.png")),
                    FileType(a.get("mimetype", "image/png")),
                    Disposition("attachment")))
            atts.extend(_sendgrid_inline_attachments(inline_photo_ids))
            message = Mail(from_email=sg_sender, to_emails=recipients,
                           subject=subject, html_content=html_body)
            message.attachment = atts
            SendGridAPIClient(sg_api_key).send(message)
        else:
            from email.mime.image import MIMEImage
            msg = MIMEMultipart("mixed")
            msg["Subject"] = subject
            msg["From"] = smtp_user
            msg["To"] = ", ".join(recipients)
            related = MIMEMultipart("related")
            related.attach(MIMEText(html_body, "html"))
            for part in _smtp_inline_photo_parts(inline_photo_ids):
                related.attach(part)
            msg.attach(related)
            for a in photo_payloads:
                img = MIMEImage(base64.b64decode(a["content_b64"]), _subtype="png")
                img.add_header("Content-Disposition", "attachment",
                               filename=a.get("filename", "photo.png"))
                msg.attach(img)
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, recipients, msg.as_string())
        return {"ok": True,
                "message": f"Service record sent to {', '.join(recipients)}",
                "count": len(records)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Job export failed: {str(e)}")


@app.post("/export/pdf-email")
async def export_pdf_email(request: Request):
    """Email a client-generated single-job PDF service record to recipients.

    Accepts multipart (file=<pdf>, recipients=<json list> or to=<email>,
    subject, body) or JSON (file_base64, filename, recipients/to, subject,
    body). Mirrors the /export/excel send logic (SendGrid or SMTP)."""
    content_type = request.headers.get("content-type", "")
    subject = "Service Record"
    body = "Service record attached."
    filename = "service-record.pdf"
    pdf_bytes = None

    if "application/json" in content_type:
        data = await request.json()
        recipients = data.get("recipients") or (
            [data["to"]] if data.get("to") else [])
        subject = (data.get("subject") or subject).strip()
        body = (data.get("body") or body)
        filename = (data.get("filename") or filename)
        b64 = data.get("file_base64")
        if b64:
            pdf_bytes = base64.b64decode(b64)
    else:
        form = await request.form()
        recipients = json.loads(form.get("recipients", "[]")) or (
            [form.get("to")] if form.get("to") else [])
        subject = (form.get("subject") or subject).strip()
        body = (form.get("body") or body)
        upload = form.get("file")
        if upload is not None and hasattr(upload, "read"):
            pdf_bytes = await upload.read()
            filename = getattr(upload, "filename", None) or filename

    recipients = [e.strip() for e in recipients if e and "@" in e]
    if not recipients:
        raise HTTPException(status_code=422, detail="At least one recipient email required")
    if not pdf_bytes:
        raise HTTPException(status_code=422, detail="PDF file is required")

    b64 = base64.b64encode(pdf_bytes).decode()

    smtp_host = os.environ.get("SMTP_HOST")
    smtp_user = os.environ.get("SMTP_USER")
    smtp_pass = os.environ.get("SMTP_PASS")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    sg_api_key = os.environ.get("SENDGRID_API_KEY")
    sg_sender = os.environ.get("SENDGRID_SENDER_EMAIL")

    if not sg_api_key and (not smtp_host or not smtp_user):
        return {"ok": True, "message": "Email not configured — file returned",
                "filename": filename, "file_base64": b64}

    try:
        if sg_api_key and SendGridAPIClient and Mail:
            from sendgrid.helpers.mail import (
                Attachment, FileContent, FileName, FileType, Disposition)
            message = Mail(from_email=sg_sender, to_emails=recipients,
                           subject=subject,
                           html_content=f"<p>{body}</p>")
            message.attachment = Attachment(
                FileContent(b64), FileName(filename),
                FileType("application/pdf"), Disposition("attachment"))
            SendGridAPIClient(sg_api_key).send(message)
        else:
            from email.mime.application import MIMEApplication
            msg = MIMEMultipart()
            msg["Subject"] = subject
            msg["From"] = smtp_user
            msg["To"] = ", ".join(recipients)
            msg.attach(MIMEText(body, "plain"))
            part = MIMEApplication(pdf_bytes, _subtype="pdf")
            part.add_header("Content-Disposition", "attachment", filename=filename)
            msg.attach(part)
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(smtp_user, recipients, msg.as_string())
        return {"ok": True, "message": f"Service record sent to {', '.join(recipients)}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")


# ─── USERS (lightweight identity for attribution) ───────────────────────────

@app.get("/users")
def list_users():
    db = SessionLocal()
    users = db.query(User).all()
    out = [{"id": u.id, "email": u.email, "name": u.name, "role": u.role} for u in users]
    db.close()
    return out


@app.post("/users")
async def create_user(data: dict = Body(...)):
    email = (data.get("email") or "").strip()
    if "@" not in email:
        raise HTTPException(status_code=422, detail="Valid email required")
    db = SessionLocal()
    existing = db.query(User).filter(User.email == email).first()
    if existing:
        out = {"id": existing.id, "email": existing.email, "name": existing.name, "role": existing.role}
        db.close()
        return out
    u = User(email=email, name=(data.get("name") or "").strip() or None,
             role=data.get("role", "field"))
    db.add(u)
    db.commit()
    db.refresh(u)
    out = {"id": u.id, "email": u.email, "name": u.name, "role": u.role}
    db.close()
    return out
